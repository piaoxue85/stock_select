#-*- encoding:utf8 -*-
import openpyxl
import tkMessageBox
from data_process import *
from pymongo import MongoClient

today_name = u'v3.2选股' + today + u'.xlsx'
today_name_future = u'v3.2期货' + today + u'.xlsx'
#居中的格式
alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center',wrap_text=True)
def open_workbook(workbook):
    try:
        wb = openpyxl.load_workbook(filename = workbook)
    except IOError,e:
        wb = openpyxl.Workbook()
    return wb
def change_sheet(wb,name):
    try:
        ws = wb.get_sheet_by_name('Sheet')
        ws.title = name
    except Exception,e:
        pass
def open_sheet(wb,sheet):
    try:
        ws = wb.get_sheet_by_name(sheet)
    except Exception,e:
        ws = wb.create_sheet()
        ws.title = sheet
    return ws
def format(ws,num,width):
    list  = ["A","B","C","D","E","F","G","H","I","J","K","L","M","N","O","P","Q","R","S","T","U","V","W","X","Y","Z"]
    for i in range(1,num):
        x = ws.column_dimensions[list[i]].width = width

#一次性执行所有
ALL = False
def write_all():
    P.exec_all()
    global ALL
    ALL = True
#公式统计
def formula():
    codes = D.all_code().split(",")
    length = len(codes)
    wb = open_workbook(today_name)
    change_sheet(wb,u'股票代码')
    ws = open_sheet(wb,u'股票代码')
    ws.cell(row=1,column=1).alignment = alignment
    ws.cell(row=1,column=2).alignment = alignment
    ws.cell(row = 1, column = 1).value = u'所有股票'
    ws.cell(row = 1, column = 2).value = u'获得星数'
    # ws.auto_filter.ref = "M1:M6"
    # ws.auto_filter.add_filter_column(1, ["Kiwi", "Apple", "Mango"])
    # ws.auto_filter.add_sort_condition('M1:M6',descending=True)
    for i in range(length):
        ws.cell(row = i+1+1, column = 1).value = str(codes[i])
        ws.cell(row = i+1, column = 1).font = openpyxl.styles.Font(color='00FFD700')
        ws.cell(row = i+1+1, column = 2).value = '=SUM(COUNTIF(INDIRECT({"C2:U'+ str(length) + '","V2:X201","Y2:Y3000"}),A' + str(i+1+1) +'))'
    wb.save(today_name)






def star():
    codes = D.Codes.split(",")
    length = len(codes)
    wb = open_workbook(today_name)
    change_sheet(wb,u'股票代码')
    ws = open_sheet(wb,u'股票代码')
    ws.cell(row=1,column=1).alignment = alignment
    ws.cell(row=1,column=2).alignment = alignment
    ws.cell(row = 1, column = 1).value = u'所有指标'
    ws.cell(row = 1, column = 2).value = u'获得星数'
    ws.cell(row=1, column=3).alignment = alignment
    ws.cell(row=1, column=4).alignment = alignment
    ws.cell(row=1, column=3).value = u'公司价值'
    ws.cell(row=1, column=4).value = u'获得星数'
    ws.cell(row=1, column=5).alignment = alignment
    ws.cell(row=1, column=6).alignment = alignment
    ws.cell(row=1, column=5).value = u'公司筹码'
    ws.cell(row=1, column=6).value = u'获得星数'
    ws.cell(row=1, column=7).alignment = alignment
    ws.cell(row=1, column=8).alignment = alignment
    ws.cell(row=1, column=7).value = u'市场波动'
    ws.cell(row=1, column=8).value = u'获得星数'
    ws.cell(row=1, column=9).alignment = alignment
    ws.cell(row=1, column=10).alignment = alignment
    ws.cell(row=1, column=9).value = u'财务'
    ws.cell(row=1, column=10).value = u'获得星数'
    all = []
    value = []
    chips = []
    market = []
    finance = []
    # 公司价值
    for y in range(num_company+1,num_company+5):
        for x in range(length):
            all.append(ws.cell(row=x+1+1,column=y).value)
            value.append(ws.cell(row=x+1+1,column=y).value)
    # 筹码
    for y in range(num_chips+1,num_chips+6):
        for x in range(length):
            all.append(ws.cell(row=x+1+1,column=y).value)
            chips.append(ws.cell(row=x+1+1,column=y).value)
    for y in range(num_chips+6, num_chips+9):
        for x in range(300):
            all.append(ws.cell(row=x + 1 + 1, column=y).value)
            chips.append(ws.cell(row=x + 1 + 1, column=y).value)
    for x in range(length):
        all.append(ws.cell(row=x + 1 + 1, column=num_chips+9).value)
        chips.append(ws.cell(row=x + 1 + 1, column=num_chips+9).value)
    for y in range(num_chips+10, num_chips+12):
        for x in range(300):
            all.append(ws.cell(row=x + 1 + 1, column=y).value)
            chips.append(ws.cell(row=x + 1 + 1, column=y).value)

    # 市场波动
    for y in range(num_stock+1, num_stock+14):
        for x in range(length):
            all.append(ws.cell(row=x + 1 + 1, column=y).value)
            market.append(ws.cell(row=x + 1 + 1, column=y).value)

    #财务
    for y in range(num_finance+1, num_finance + 4):
        for x in range(300):
            all.append(ws.cell(row=x + 1 + 1, column=y).value)
            finance.append(ws.cell(row=x + 1 + 1, column=y).value)
    # 总共
    myset = set(all)
    myset_value = set(value)
    myset_chips = set(chips)
    myset_market = set(market)
    myset_finance = set(finance)
    result_all = []
    result_value = []
    result_chips = []
    result_market = []
    result_finance = []
    for each in myset:
        result_all.append((each,all.count(each)))
    for each in myset_value:
        result_value.append((each, value.count(each)))
    for each in myset_chips:
        result_chips.append((each, chips.count(each)))
    for each in myset_market:
        result_market.append((each, market.count(each)))
    for each in myset_finance:
        result_finance.append((each, finance.count(each)))
    result_all_sort = sorted(result_all,key = lambda result: result[1],reverse=True)
    result_value_sort = sorted(result_value,key = lambda result: result[1],reverse=True)
    result_chips_sort = sorted(result_chips,key = lambda result: result[1],reverse=True)
    result_market_sort = sorted(result_market,key = lambda result: result[1],reverse=True)
    result_finance_sort = sorted(result_finance,key = lambda result: result[1],reverse=True)
    for i in range(1,len(result_all)):
        ws.cell(row = i+1, column = 1).value = result_all_sort[i][0]
        ws.cell(row = i+1, column = 1).font = openpyxl.styles.Font(color='00000000')
        ws.cell(row = i+1, column = 2).value = result_all_sort[i][1]
    for i in range(1, len(result_value)):
        ws.cell(row=i + 1, column=3).value = result_value_sort[i][0]
        ws.cell(row=i + 1, column=3).font = openpyxl.styles.Font(color='00000000')
        ws.cell(row=i + 1, column=4).value = result_value_sort[i][1]
    for i in range(1, len(result_chips)):
        ws.cell(row=i + 1, column=5).value = result_chips_sort[i][0]
        ws.cell(row=i + 1, column=5).font = openpyxl.styles.Font(color='00000000')
        ws.cell(row=i + 1, column=6).value = result_chips_sort[i][1]
    for i in range(1, len(result_market)):
        ws.cell(row=i + 1, column=7).value = result_market_sort[i][0]
        ws.cell(row=i + 1, column=7).font = openpyxl.styles.Font(color='00000000')
        ws.cell(row=i + 1, column=8).value = result_market_sort[i][1]
    for i in range(1, len(result_finance)):
        ws.cell(row=i + 1, column=9).value = result_finance_sort[i][0]
        ws.cell(row=i + 1, column=9).font = openpyxl.styles.Font(color='00000000')
        ws.cell(row=i + 1, column=10).value = result_finance_sort[i][1]
    wb.save(today_name)

# position代表了该指标在总表的列数，方便修改
# 公司价值数量
num_company = 10
# 筹码
num_chips = 10+4
# 市场波动
num_stock = 10+4+11
#财务
num_finance = 10+4+11+13
# 1 #营收季增率
position1 = num_company+1
# 2 #EPS季增率
position2 = num_company+2
# 3 #毛利率季增率
position3 = num_company+3
# 4 #净资产收益率ROE净增率
position4 = num_company+4
# 5 #波动率今10MA > 前10MA，同时 10 MA > 100MA
position5 = num_stock+1
# 6 #价格今10MA > 前10MA，同时前10MA > 前前10MA
position6 = num_stock+2
# 7 #波动性:日线 周线 月线10MA都上扬
position7 = num_stock+3
# 8 #日线10MA上扬
position8 = num_stock+4
# 9 #周线10MA上扬
position9 = num_stock+5
# 10 #月线10MA上扬
position10 = num_stock+6
# 11 #60 180 250天波动区间在30% 50% 100%内
position11 = num_stock+7
# 12 #60天波动区间在30%内
position12 = num_stock+8
# 13 #180天波动区间在50%内
position13 = num_stock+9
# 14 #250天波动区间在100%内
position14 = num_stock+10
# 15 #融资余额今10MA>前10MA,同时10MA>30MA
position15 = num_chips+1
# 16 #户均持股前三季中有一次增加
position16 = num_chips+2
# 17 #股东户数两季连续减少
position17 = num_chips+3
# 18 #十大股东变化超过3个
position18 = num_chips+4
# 19 #十大股东持股比例增加
position19 = num_chips+5
# 20、21、22 #券商、基金、机构持股数量排序
position20 = num_chips+6
position21 = num_chips+7
position22 = num_chips+8
# 23 #换手率5MA大于前一天5MA，同时10MA大于前一天10MA
position23 = num_chips+9
# 24 #涨幅日、周、月都超过大盘
position24 = num_stock+11
# 25 #袁氏选股
position25 = num_stock+12
position26 = num_stock+13
# 27 # 股本，总股本和流通股本
position27 = num_chips+10
position28 = num_chips+11

# 29 # 公司净值（每股净资产BPS）
position29 = num_finance+1
# 30 # 市盈率PE
position30 = num_finance+2
# 31 # 每股分红送转
position31 = num_finance+3

# 1 #营收季增率


def strategy1():
    revenue = P.process1()
    wb = open_workbook(today_name)
    change_sheet(wb,u'股票代码')
    ws = open_sheet(wb,u'股票代码')
    ft = openpyxl.styles.Font(color=openpyxl.styles.colors.RED)
    ws.cell(row = 1,column = position1).value = u'营收季增率'
    ws.cell(row = 1,column = position1).alignment = alignment
    for i in range(len(revenue)):
        ws.cell(row = i+1+1,column = position1).value = revenue[i][0]
        ws.cell(row = i+1+1,column = position1).font = openpyxl.styles.Font(color=openpyxl.styles.colors.BLUE)
    ws = open_sheet(wb,u'营收季增率')
    ws.cell(row=1,column=1).value = u'符合条件代码'
    ws.cell(row=1,column=2).value = u'上三季营业收入'+'\n'+ revenue[0][4][0].isoformat()[0:10]
    ws.cell(row=1,column=3).value = u'上两季营业收入'+'\n'+ revenue[0][4][1].isoformat()[0:10]
    ws.cell(row=1,column=4).value = u'上一季营业收入'+'\n'+ revenue[0][4][2].isoformat()[0:10]
    ws.cell(row=1,column=5).value = u'上两季比上三季增长率'
    ws.cell(row=1,column=6).value = u'上一季比上两季增长率'
    ws.cell(row=1,column=1).alignment = alignment
    ws.cell(row=1,column=2).alignment = alignment
    ws.cell(row=1,column=3).alignment = alignment
    ws.cell(row=1,column=4).alignment = alignment
    ws.cell(row=1,column=5).alignment = alignment
    ws.cell(row=1,column=6).alignment = alignment
    format(ws,6,20.0)
    for i in range(0,len(revenue)):
        ws.cell(row = i+1+1, column = 2).number_format = '#,##0.00_);[Green]-#,##0.00'
        ws.cell(row = i+1+1, column = 3).number_format = '#,##0.00_);[Green]-#,##0.00'
        ws.cell(row = i+1+1, column = 4).number_format = '#,##0.00_);[Green]-#,##0.00'
        ws.cell(row = i+1+1, column = 5).number_format = '0.00%;[Green]-0.00%'
        ws.cell(row = i+1+1, column = 6).number_format = '0.00%;[Green]-0.00%'
        ws.cell(row = i+1+1, column = 5).font = ft
        ws.cell(row = i+1+1, column = 6).font = ft
        ws.cell(row = i+1+1, column = 1).value = revenue[i][0]
        ws.cell(row = i+1+1, column = 2).value = revenue[i][3][0]
        ws.cell(row = i+1+1, column = 3).value = revenue[i][3][1]
        ws.cell(row = i+1+1, column = 4).value = revenue[i][3][2]
        ws.cell(row = i+1+1, column = 5).value = revenue[i][2]
        ws.cell(row = i+1+1, column = 6).value = revenue[i][1]
    wb.save(today_name)
    # if(ALL == False):
    #     tkMessageBox.showinfo(title=u'写入成功', message=u'策略一已经成功，请点击打开进行浏览')
# 2 #EPS季增率


def strategy2():
    EPS = P.process2()
    wb = open_workbook(today_name)
    change_sheet(wb,u'股票代码')
    ws = open_sheet(wb,u'股票代码')
    ft = openpyxl.styles.Font(color=openpyxl.styles.colors.RED)
    ws.cell(row = 1,column = position2).value = u'EPS季增率'
    ws.cell(row = 1,column = position2).alignment = alignment
    for i in range(len(EPS)):
        ws.cell(row = i+1+1,column = position2).value = EPS[i][0]
        ws.cell(row = i+1+1,column = position2).font = openpyxl.styles.Font(color=openpyxl.styles.colors.BLUE)
    ws = open_sheet(wb,u'EPS季增率')
    ws.cell(row=1,column=1).value = u'符合条件代码'
    ws.cell(row=1,column=2).value = u'上三季EPS'+'\n'+ EPS[0][4][0].isoformat()[0:10]
    ws.cell(row=1,column=3).value = u'上两季EPS'+'\n'+ EPS[0][4][1].isoformat()[0:10]
    ws.cell(row=1,column=4).value = u'上一季EPS'+'\n'+ EPS[0][4][2].isoformat()[0:10]
    ws.cell(row=1,column=5).value = u'上两季比上三季增长率'
    ws.cell(row=1,column=6).value = u'上一季比上两季增长率'
    ws.cell(row=1,column=1).alignment = alignment
    ws.cell(row=1,column=2).alignment = alignment
    ws.cell(row=1,column=3).alignment = alignment
    ws.cell(row=1,column=4).alignment = alignment
    ws.cell(row=1,column=5).alignment = alignment
    ws.cell(row=1,column=6).alignment = alignment
    format(ws,6,20.0)
    for i in range(0,len(EPS)):
        ws.cell(row = i+1+1, column = 1).font = openpyxl.styles.Font(color=openpyxl.styles.colors.BLUE)
        ws.cell(row = i+1+1, column = 2).number_format = '0.00%;[Green]-0.00%'
        ws.cell(row = i+1+1, column = 3).number_format = '0.00%;[Green]-0.00%'
        ws.cell(row = i+1+1, column = 4).number_format = '0.00%;[Green]-0.00%'
        ws.cell(row = i+1+1, column = 5).number_format = '0.00%;[Green]-0.00%'
        ws.cell(row = i+1+1, column = 6).number_format = '0.00%;[Green]-0.00%'
        ws.cell(row = i+1+1, column = 5).font = ft
        ws.cell(row = i+1+1, column = 6).font = ft
        ws.cell(row = i+1+1, column = 1).value = EPS[i][0]
        ws.cell(row = i+1+1, column = 2).value = EPS[i][3][0]
        ws.cell(row = i+1+1, column = 3).value = EPS[i][3][1]
        ws.cell(row = i+1+1, column = 4).value = EPS[i][3][2]
        ws.cell(row = i+1+1, column = 5).value = EPS[i][2]
        ws.cell(row = i+1+1, column = 6).value = EPS[i][1]
    wb.save(today_name)
    # if(ALL == False):
    #     tkMessageBox.showinfo(title=u'写入成功', message=u'策略二已经成功，请点击打开进行浏览')

# 3 #毛利率季增率


def strategy3():
    MLL = P.process3()
    wb = open_workbook(today_name)
    change_sheet(wb,u'股票代码')
    ws = open_sheet(wb,u'股票代码')
    ft = openpyxl.styles.Font(color=openpyxl.styles.colors.RED)
    ws.cell(row = 1,column = position3).value = u'毛利率季增率'
    ws.cell(row = 1,column = position3).alignment = alignment
    for i in range(len(MLL)):
        ws.cell(row = i+1+1,column = position3).value = MLL[i][0]
        ws.cell(row = i+1+1,column = position3).font = openpyxl.styles.Font(color=openpyxl.styles.colors.BLUE)
    ws = open_sheet(wb,u'毛利率季增率')
    ws.cell(row=1,column=1).value = u'符合条件代码'
    ws.cell(row=1,column=2).value = u'上三季毛利率'+'\n'+ MLL[0][4][0].isoformat()[0:10]
    ws.cell(row=1,column=3).value = u'上两季毛利率'+'\n'+ MLL[0][4][1].isoformat()[0:10]
    ws.cell(row=1,column=4).value = u'上一季毛利率'+'\n'+ MLL[0][4][2].isoformat()[0:10]
    ws.cell(row=1,column=5).value = u'上两季比上三季增长率'
    ws.cell(row=1,column=6).value = u'上一季比上两季增长率'
    ws.cell(row=1,column=1).alignment = alignment
    ws.cell(row=1,column=2).alignment = alignment
    ws.cell(row=1,column=3).alignment = alignment
    ws.cell(row=1,column=4).alignment = alignment
    ws.cell(row=1,column=5).alignment = alignment
    ws.cell(row=1,column=6).alignment = alignment
    format(ws,6,20.0)
    for i in range(0,len(MLL)):
        ws.cell(row = i+1+1, column = 1).font = openpyxl.styles.Font(color=openpyxl.styles.colors.BLUE)
        ws.cell(row = i+1+1, column = 2).number_format = '#,##0.00_)\%;[Green]-#,##0.00\%'
        ws.cell(row = i+1+1, column = 3).number_format = '#,##0.00_)\%;[Green]-#,##0.00\%'
        ws.cell(row = i+1+1, column = 4).number_format = '#,##0.00_)\%;[Green]-#,##0.00\%'
        ws.cell(row = i+1+1, column = 5).number_format = '0.00%;[Green]-0.00%'
        ws.cell(row = i+1+1, column = 6).number_format = '0.00%;[Green]-0.00%'
        ws.cell(row = i+1+1, column = 5).font = ft
        ws.cell(row = i+1+1, column = 6).font = ft
        ws.cell(row = i+1+1, column = 1).value = MLL[i][0]
        ws.cell(row = i+1+1, column = 2).value = MLL[i][3][0]
        ws.cell(row = i+1+1, column = 3).value = MLL[i][3][1]
        ws.cell(row = i+1+1, column = 4).value = MLL[i][3][2]
        ws.cell(row = i+1+1, column = 5).value = MLL[i][2]
        ws.cell(row = i+1+1, column = 6).value = MLL[i][1]
    wb.save(today_name)
    # if(ALL == False):
    #     tkMessageBox.showinfo(title=u'写入成功', message=u'策略三已经成功，请点击打开进行浏览')

# 4 #净资产收益率ROE净增率


def strategy4():
    ROE = P.process4()
    wb = open_workbook(today_name)
    change_sheet(wb,u'股票代码')
    ws = open_sheet(wb,u'股票代码')
    ft = openpyxl.styles.Font(color=openpyxl.styles.colors.RED)
    ws.cell(row = 1,column = position4).value = u'ROE季增率'
    ws.cell(row = 1,column = position4).alignment = alignment
    for i in range(len(ROE)):
        ws.cell(row = i+1+1,column = position4).value = ROE[i][0]
        ws.cell(row = i+1+1,column = position4).font = openpyxl.styles.Font(color=openpyxl.styles.colors.BLUE)
    ws = open_sheet(wb,u'ROE季增率')
    ws.cell(row=1,column=1).value = u'符合条件代码'
    ws.cell(row=1,column=2).value = u'上三季净资产收益率'+'\n'+ ROE[0][4][0].isoformat()[0:10]
    ws.cell(row=1,column=3).value = u'上两季净资产收益率'+'\n'+ ROE[0][4][1].isoformat()[0:10]
    ws.cell(row=1,column=4).value = u'上一季净资产收益率'+'\n'+ ROE[0][4][2].isoformat()[0:10]
    ws.cell(row=1,column=5).value = u'上两季比上三季增长率'
    ws.cell(row=1,column=6).value = u'上一季比上两季增长率'
    ws.cell(row=1,column=1).alignment = alignment
    ws.cell(row=1,column=2).alignment = alignment
    ws.cell(row=1,column=3).alignment = alignment
    ws.cell(row=1,column=4).alignment = alignment
    ws.cell(row=1,column=5).alignment = alignment
    ws.cell(row=1,column=6).alignment = alignment
    format(ws,6,20.0)
    for i in range(0,len(ROE)):
        ws.cell(row = i+1+1, column = 1).font = openpyxl.styles.Font(color=openpyxl.styles.colors.BLUE)
        ws.cell(row = i+1+1, column = 2).number_format = '#,##0.00_)\%;[Green]-#,##0.00\%'
        ws.cell(row = i+1+1, column = 3).number_format = '#,##0.00_)\%;[Green]-#,##0.00\%'
        ws.cell(row = i+1+1, column = 4).number_format = '#,##0.00_)\%;[Green]-#,##0.00\%'
        ws.cell(row = i+1+1, column = 5).number_format = '0.00%;[Green]-0.00%'
        ws.cell(row = i+1+1, column = 6).number_format = '0.00%;[Green]-0.00%'
        ws.cell(row = i+1+1, column = 5).font = ft
        ws.cell(row = i+1+1, column = 6).font = ft
        ws.cell(row = i+1+1, column = 1).value = ROE[i][0]
        ws.cell(row = i+1+1, column = 2).value = ROE[i][3][0]
        ws.cell(row = i+1+1, column = 3).value = ROE[i][3][1]
        ws.cell(row = i+1+1, column = 4).value = ROE[i][3][2]
        ws.cell(row = i+1+1, column = 5).value = ROE[i][2]
        ws.cell(row = i+1+1, column = 6).value = ROE[i][1]
    wb.save(today_name)
    # if(ALL == False):
    #     tkMessageBox.showinfo(title=u'写入成功', message=u'策略四已经成功，请点击打开进行浏览')

# 5 #波动率今10MA > 前10MA，同时 10 MA > 100MA


def strategy5():
    volatility = P.process5()
    wb = open_workbook(today_name)
    change_sheet(wb,u'股票代码')
    ws = open_sheet(wb,u'股票代码')
    ft = openpyxl.styles.Font(color=openpyxl.styles.colors.RED)
    ws.cell(row = 1,column = position5).value = u'波动率'
    ws.cell(row = 1,column = position5).alignment = alignment
    for i in range(len(volatility)):
        ws.cell(row = i+1+1,column = position5).value = volatility[i][0]
        ws.cell(row = i+1+1,column = position5).font = openpyxl.styles.Font(color=openpyxl.styles.colors.RED)
    ws = open_sheet(wb,u'波动率')
    ws.cell(row = 1, column = 1).value = u'符合条件代码'
    ws.cell(row = 1, column = 2).value = u'今日波动率\n(10天为周期)\n' + volatility[0][2][110].isoformat()[0:10]
    ws.cell(row = 1, column = 3).value = u'前一天10MA'
    ws.cell(row = 1, column = 4).value = u'今日10MA'
    ws.cell(row = 1, column = 5).value = u'100MA'
    ws.cell(row=1,column=1).alignment = alignment
    ws.cell(row=1,column=2).alignment = alignment
    ws.cell(row=1,column=3).alignment = alignment
    ws.cell(row=1,column=4).alignment = alignment
    ws.cell(row=1,column=5).alignment = alignment
    format(ws,5,15.0)
    for i in range(0,len(volatility)):
        ws.cell(row = i+1+1,column = 1).font = openpyxl.styles.Font(color=openpyxl.styles.colors.BLUE)
        ws.cell(row = i+1+1, column = 2).number_format = '0.00000%;[Green]-0.00000%'
        ws.cell(row = i+1+1, column = 3).number_format = '0.00000%;[Green]-0.00000%'
        ws.cell(row = i+1+1, column = 4).number_format = '0.00000%;[Green]-0.00000%'
        ws.cell(row = i+1+1, column = 5).number_format = '0.00000%;[Green]-0.00000%'
        ws.cell(row = i+1+1, column = 1).value = volatility[i][0]
        ws.cell(row = i+1+1, column = 2).value = volatility[i][1][0]
        ws.cell(row = i+1+1, column = 3).value = volatility[i][1][1]
        ws.cell(row = i+1+1, column = 4).value = volatility[i][1][2]
        ws.cell(row = i+1+1, column = 5).value = volatility[i][1][3]
    wb.save(today_name)
    # if(ALL == False):
    #     tkMessageBox.showinfo(title=u'写入成功', message=u'策略五已经成功，请点击打开进行浏览')

# 6 #价格今10MA > 前10MA，同时前10MA > 前前10MA


def strategy6():
    price = P.process6()
    wb = open_workbook(today_name)
    change_sheet(wb,u'股票代码')
    ws = open_sheet(wb,u'股票代码')
    ws.cell(row = 1,column = position6).value = u'价格MA'
    ws.cell(row = 1,column = position6).alignment = alignment
    for i in range(len(price)):
        ws.cell(row = i+1+1,column = position6).value = price[i][0]
        ws.cell(row = i+1+1,column = position6).font = openpyxl.styles.Font(color=openpyxl.styles.colors.RED)
    ws = open_sheet(wb,u'价格MA')
    ws.cell(row = 1, column = 1).value = u'符合条件代码'
    ws.cell(row = 1, column = 2).value = u'前两天10MA\n' + price[0][2][9].isoformat()[0:10]
    ws.cell(row = 1, column = 3).value = u'前一天10MA\n' + price[0][2][10].isoformat()[0:10]
    ws.cell(row = 1, column = 4).value = u'今日10MA\n' + price[0][2][11].isoformat()[0:10]
    ws.cell(row=1,column=1).alignment = alignment
    ws.cell(row=1,column=2).alignment = alignment
    ws.cell(row=1,column=3).alignment = alignment
    ws.cell(row=1,column=4).alignment = alignment
    format(ws,4,15.0)
    for i in range(0,len(price)):
        ws.cell(row = i+1+1,column = 1).font = openpyxl.styles.Font(color=openpyxl.styles.colors.BLUE)
        ws.cell(row = i+1+1, column = 2).number_format = '#,##0.00_);[Green]-#,##0.00'
        ws.cell(row = i+1+1, column = 3).number_format = '#,##0.00_);[Green]-#,##0.00'
        ws.cell(row = i+1+1, column = 4).number_format = '#,##0.00_);[Green]-#,##0.00'
        ws.cell(row = i+1+1,column = 1).value = price[i][0]
        ws.cell(row = i+1+1, column = 2).value = price[i][1][0]
        ws.cell(row = i+1+1, column = 3).value = price[i][1][1]
        ws.cell(row = i+1+1, column = 4).value = price[i][1][2]
    wb.save(today_name)
    # if(ALL == False):
    #     tkMessageBox.showinfo(title=u'写入成功', message=u'策略六已经成功，请点击打开进行浏览')

# 7 #波动性:日线 周线 月线10MA都上扬


def strategy7():
    synchronization = P.process7()
    wb = open_workbook(today_name)
    change_sheet(wb,u'股票代码')
    ws = open_sheet(wb,u'股票代码')
    if len(synchronization):
        last_day = synchronization[0][2][9].isoformat()[0:10]
        this_day = synchronization[0][2][10].isoformat()[0:10]
        last_week = synchronization[0][3][9].isoformat()[0:10]
        this_week = synchronization[0][3][10].isoformat()[0:10]
        last_month = synchronization[0][4][9].isoformat()[0:10]
        this_month = synchronization[0][4][10].isoformat()[0:10]
    else:
        last_day = str(' ')
        this_day = str(' ')
        last_week = str(' ')
        this_week = str(' ')
        last_month = str(' ')
        this_month = str(' ')
    ws.cell(row = 1,column = position7).value = u'日周月10MA'
    ws.cell(row = 1,column = position7).alignment = alignment
    for i in range(len(synchronization)):
        ws.cell(row = i+1+1,column = position7).value = synchronization[i][0]
        ws.cell(row = i+1+1,column = position7).font = openpyxl.styles.Font(color=openpyxl.styles.colors.RED)
    ws = open_sheet(wb,u'日周月10MA')
    ws.cell(row = 1, column = 1).value = u'符合条件代码'
    ws.cell(row = 1, column = 2).value = u'前一天10MA日线\n' + last_day
    ws.cell(row = 1, column = 3).value = u'10MA日线\n' + this_day
    ws.cell(row = 1, column = 4).value = u'前一周10MA周线\n' + last_week
    ws.cell(row = 1, column = 5).value = u'10MA周线\n' + this_week
    ws.cell(row = 1, column = 6).value = u'前一月10MA月线\n' + last_month
    ws.cell(row = 1, column = 7).value = u'10MA月线\n' + this_month
    ws.cell(row=1,column=1).alignment = alignment
    ws.cell(row=1,column=2).alignment = alignment
    ws.cell(row=1,column=3).alignment = alignment
    ws.cell(row=1,column=4).alignment = alignment
    ws.cell(row=1,column=5).alignment = alignment
    ws.cell(row=1,column=6).alignment = alignment
    ws.cell(row=1,column=7).alignment = alignment
    format(ws,7,17.0)
    for i in range(0,len(synchronization)):
        ws.cell(row = i+1+1,column = 1).font = openpyxl.styles.Font(color=openpyxl.styles.colors.BLUE)
        ws.cell(row = i+1+1,column = 1).value = synchronization[i][0]
        ws.cell(row = i+1+1, column = 2).value = synchronization[i][1][0]
        ws.cell(row = i+1+1, column = 3).value = synchronization[i][1][1]
        ws.cell(row = i+1+1, column = 4).value = synchronization[i][1][2]
        ws.cell(row = i+1+1, column = 5).value = synchronization[i][1][3]
        ws.cell(row = i+1+1, column = 6).value = synchronization[i][1][4]
        ws.cell(row = i+1+1, column = 7).value = synchronization[i][1][5]
    wb.save(today_name)
    # if(ALL == False):
    #     tkMessageBox.showinfo(title=u'写入成功', message=u'策略七已经成功，请点击打开进行浏览')

# 8 #日线10MA上扬


def strategy8():
    DAY10 = P.process8()
    wb = open_workbook(today_name)
    change_sheet(wb,u'股票代码')
    ws = open_sheet(wb,u'股票代码')
    ws.cell(row = 1,column = position8).value = u'日10MA'
    ws.cell(row = 1,column = position8).alignment = alignment
    for i in range(len(DAY10)):
        ws.cell(row = i+1+1,column = position8).value = DAY10[i][0]
        ws.cell(row = i+1+1,column = position8).font = openpyxl.styles.Font(color=openpyxl.styles.colors.RED)
    ws = open_sheet(wb,u'日10MA')
    ws.cell(row = 1, column = 1).value = u'符合条件代码'
    ws.cell(row = 1, column = 2).value = u'前一天10MA日线\n' + DAY10[0][2][9].isoformat()[0:10]
    ws.cell(row = 1, column = 3).value = u'10MA日线\n' + DAY10[0][2][10].isoformat()[0:10]
    ws.cell(row=1,column=1).alignment = alignment
    ws.cell(row=1,column=2).alignment = alignment
    ws.cell(row=1,column=3).alignment = alignment
    format(ws,4,17.0)
    for i in range(0,len(DAY10)):
        ws.cell(row = i+1+1,column = 1).font = openpyxl.styles.Font(color=openpyxl.styles.colors.BLUE)
        ws.cell(row = i+1+1,column = 1).value = DAY10[i][0]
        ws.cell(row = i+1+1, column = 2).value = DAY10[i][1][0]
        ws.cell(row = i+1+1, column = 3).value = DAY10[i][1][1]
    wb.save(today_name)
    # if(ALL == False):
    #     tkMessageBox.showinfo(title=u'写入成功', message=u'策略八已经成功，请点击打开进行浏览')

# 9 #周线10MA上扬


def strategy9():
    WEEK10 = P.process9()
    if len(WEEK10):
        last_day = WEEK10[0][2][9].isoformat()[0:10]
        MA10 = WEEK10[0][2][10].isoformat()[0:10]
    else:
        last_day = str(' ')
        MA10= str(' ')
    wb = open_workbook(today_name)
    change_sheet(wb,u'股票代码')
    ws = open_sheet(wb,u'股票代码')
    ws.cell(row = 1,column = position9).value = u'周10MA'
    ws.cell(row = 1,column = position9).alignment = alignment
    for i in range(len(WEEK10)):
        ws.cell(row = i+1+1,column = position9).value = WEEK10[i][0]
        ws.cell(row = i+1+1,column = position9).font = openpyxl.styles.Font(color=openpyxl.styles.colors.RED)
    ws = open_sheet(wb,u'周10MA')
    ws.cell(row = 1, column = 1).value = u'符合条件代码'
    ws.cell(row = 1, column = 2).value = u'前一天10MA周线\n' + last_day
    ws.cell(row = 1, column = 3).value = u'10MA周线\n' + MA10
    ws.cell(row=1,column=1).alignment = alignment
    ws.cell(row=1,column=2).alignment = alignment
    ws.cell(row=1,column=3).alignment = alignment
    format(ws,4,17.0)
    for i in range(0,len(WEEK10)):
        ws.cell(row = i+1+1,column = 1).font = openpyxl.styles.Font(color=openpyxl.styles.colors.BLUE)
        ws.cell(row = i+1+1,column = 1).value = WEEK10[i][0]
        ws.cell(row = i+1+1, column = 2).value = WEEK10[i][1][0]
        ws.cell(row = i+1+1, column = 3).value = WEEK10[i][1][1]
    wb.save(today_name)
    # if(ALL == False):
    #     tkMessageBox.showinfo(title=u'写入成功', message=u'策略九已经成功，请点击打开进行浏览')

# 10 #月线10MA上扬


def strategy10():
    MON10 = P.process10()

    if len(MON10):
        last_month = MON10[0][2][9].isoformat()[0:10]
        this_month = MON10[0][2][10].isoformat()[0:10]
    else:
        last_month = str(' ')
        this_month = str(' ')
    wb = open_workbook(today_name)
    change_sheet(wb,u'股票代码')
    ws = open_sheet(wb,u'股票代码')
    ws.cell(row = 1,column = position10).value = u'月10MA'
    ws.cell(row = 1,column = position10).alignment = alignment
    for i in range(len(MON10)):
        ws.cell(row = i+1+1,column = position10).value = MON10[i][0]
        ws.cell(row = i+1+1,column = position10).font = openpyxl.styles.Font(color=openpyxl.styles.colors.RED)
    ws = open_sheet(wb,u'月10MA')
    ws.cell(row = 1, column = 1).value = u'符合条件代码'
    ws.cell(row = 1, column = 2).value = u'前一天10MA月线\n' + last_month
    ws.cell(row = 1, column = 3).value = u'10MA月线\n' + this_month
    ws.cell(row=1,column=1).alignment = alignment
    ws.cell(row=1,column=2).alignment = alignment
    ws.cell(row=1,column=3).alignment = alignment
    format(ws,4,17.0)
    for i in range(0,len(MON10)):
        ws.cell(row = i+1+1,column = 1).font = openpyxl.styles.Font(color=openpyxl.styles.colors.BLUE)
        ws.cell(row = i+1+1,column = 1).value = MON10[i][0]
        ws.cell(row = i+1+1, column = 2).value = MON10[i][1][0]
        ws.cell(row = i+1+1, column = 3).value = MON10[i][1][1]
    wb.save(today_name)
    # if(ALL == False):
    #     tkMessageBox.showinfo(title=u'写入成功', message=u'策略十已经成功，请点击打开进行浏览')

# 11 #60 180 250天波动区间在30% 50% 100%内


def strategy11():
    vola_range = P.process11()
    wb = open_workbook(today_name)
    change_sheet(wb,u'股票代码')
    ws = open_sheet(wb,u'股票代码')
    ws.cell(row = 1,column = position11).value = u'波动区间内'
    ws.cell(row = 1,column = position11).alignment = alignment
    for i in range(len(vola_range)):
        ws.cell(row = i+1+1,column = position11).value = vola_range[i][0]
        ws.cell(row = i+1+1,column = position11).font = openpyxl.styles.Font(color=openpyxl.styles.colors.RED)
    ws = open_sheet(wb,u'波动区间内')
    ws.cell(row = 1, column = 1).value = u'符合条件代码'
    ws.cell(row = 1, column = 2).value = u'60日内最高价\n('+ vola_range[0][2][0] + u'至今)'
    ws.cell(row = 1, column = 3).value = u'60日内最低价\n('+ vola_range[0][2][0] + u'至今)'
    ws.cell(row = 1, column = 4).value = u'180日内最高价\n('+ vola_range[0][2][1] + u'至今)'
    ws.cell(row = 1, column = 5).value = u'180日内最低价\n('+ vola_range[0][2][1] + u'至今)'
    ws.cell(row = 1, column = 6).value = u'250日内最高价\n('+ vola_range[0][2][2] + u'至今)'
    ws.cell(row = 1, column = 7).value = u'250日内最低价\n('+ vola_range[0][2][2] + u'至今)'
    ws.cell(row=1,column=1).alignment = alignment
    ws.cell(row=1,column=2).alignment = alignment
    ws.cell(row=1,column=3).alignment = alignment
    ws.cell(row=1,column=4).alignment = alignment
    ws.cell(row=1,column=5).alignment = alignment
    ws.cell(row=1,column=6).alignment = alignment
    ws.cell(row=1,column=7).alignment = alignment
    format(ws,7,18.0)
    for i in range(0,len(vola_range)):
        ws.cell(row = i+1+1,column = 1).font = openpyxl.styles.Font(color=openpyxl.styles.colors.BLUE)
        ws.cell(row = i+1+1,column = 1).value = vola_range[i][0]
        ws.cell(row = i+1+1, column = 2).value = vola_range[i][1][0]
        ws.cell(row = i+1+1, column = 3).value = vola_range[i][1][1]
        ws.cell(row = i+1+1, column = 4).value = vola_range[i][1][2]
        ws.cell(row = i+1+1, column = 5).value = vola_range[i][1][3]
        ws.cell(row = i+1+1, column = 6).value = vola_range[i][1][4]
        ws.cell(row = i+1+1, column = 7).value = vola_range[i][1][5]
    wb.save(today_name)
    # if(ALL == False):
    #     tkMessageBox.showinfo(title=u'写入成功', message=u'策略十一已经成功，请点击打开进行浏览')

# 12 #60天波动区间在30%内


def strategy12():
    vola_range60 = P.process12()
    wb = open_workbook(today_name)
    change_sheet(wb,u'股票代码')
    ws = open_sheet(wb,u'股票代码')
    ws.cell(row = 1,column = position12).value = u'60日波动小于30%'
    ws.cell(row = 1,column = position12).alignment = alignment
    for i in range(len(vola_range60)):
        ws.cell(row = i+1+1,column = position12).value = vola_range60[i][0]
        ws.cell(row = i+1+1,column = position12).font = openpyxl.styles.Font(color=openpyxl.styles.colors.RED)
    ws = open_sheet(wb,u'60日波动小于30%')
    ws.cell(row = 1, column = 1).value = u'符合条件代码'
    ws.cell(row = 1, column = 2).value = u'60日内最高价\n（'+ vola_range60[0][2] + u'至今)'
    ws.cell(row = 1, column = 3).value = u'60日内最低价\n（'+ vola_range60[0][2] + u'至今)'
    ws.cell(row=1,column=1).alignment = alignment
    ws.cell(row=1,column=2).alignment = alignment
    ws.cell(row=1,column=3).alignment = alignment
    format(ws,4,18.0)
    for i in range(0,len(vola_range60)):
        ws.cell(row = i+1+1,column = 1).font = openpyxl.styles.Font(color=openpyxl.styles.colors.BLUE)
        ws.cell(row = i+1+1,column = 1).value = vola_range60[i][0]
        ws.cell(row = i+1+1, column = 2).value = vola_range60[i][1][0]
        ws.cell(row = i+1+1, column = 3).value = vola_range60[i][1][1]
    wb.save(today_name)
    # if(ALL == False):
    #     tkMessageBox.showinfo(title=u'写入成功', message=u'策略十二已经成功，请点击打开进行浏览')

# 13 #180天波动区间在50%内


def strategy13():
    vola_range180 = P.process13()
    wb = open_workbook(today_name)
    change_sheet(wb,u'股票代码')
    ws = open_sheet(wb,u'股票代码')
    ws.cell(row = 1,column = position13).value = u'180日波动小于50%'
    ws.cell(row = 1,column = position13).alignment = alignment
    for i in range(len(vola_range180)):
        ws.cell(row = i+1+1,column = position13).value = vola_range180[i][0]
        ws.cell(row = i+1+1,column = position13).font = openpyxl.styles.Font(color=openpyxl.styles.colors.RED)
    ws = open_sheet(wb,u'180日波动小于50%')
    ws.cell(row = 1, column = 1).value = u'符合条件代码'
    ws.cell(row = 1, column = 2).value = u'180日内最高价\n（'+ vola_range180[0][2] + u'至今）'
    ws.cell(row = 1, column = 3).value = u'180日内最低价\n（'+ vola_range180[0][2] + u'至今）'
    ws.cell(row=1,column=1).alignment = alignment
    ws.cell(row=1,column=2).alignment = alignment
    ws.cell(row=1,column=3).alignment = alignment
    format(ws,4,18.0)
    for i in range(0,len(vola_range180)):
        ws.cell(row = i+1+1,column = 1).font = openpyxl.styles.Font(color=openpyxl.styles.colors.BLUE)
        ws.cell(row = i+1+1,column = 1).value = vola_range180[i][0]
        ws.cell(row = i+1+1, column = 2).value = vola_range180[i][1][0]
        ws.cell(row = i+1+1, column = 3).value = vola_range180[i][1][1]
    wb.save(today_name)
    # if(ALL == False):
    #     tkMessageBox.showinfo(title=u'写入成功', message=u'策略十三已经成功，请点击打开进行浏览')

# 14 #250天波动区间在100%内


def strategy14():
    vola_range250 = P.process14()
    wb = open_workbook(today_name)
    change_sheet(wb,u'股票代码')
    ws = open_sheet(wb,u'股票代码')
    ws.cell(row = 1,column = position14).value = u'250日波动小于100%'
    ws.cell(row = 1,column = position14).alignment = alignment
    for i in range(len(vola_range250)):
        ws.cell(row = i+1+1,column = position14).value = vola_range250[i][0]
        ws.cell(row = i+1+1,column = position14).font = openpyxl.styles.Font(color=openpyxl.styles.colors.RED)
    ws = open_sheet(wb,u'250日波动小于100%')
    ws.cell(row = 1, column = 1).value = u'符合条件代码'
    ws.cell(row = 1, column = 2).value = u'250日内最高价\n（'+ vola_range250[0][2] + u'至今）'
    ws.cell(row = 1, column = 3).value = u'250日内最低价\n（'+ vola_range250[0][2] + u'至今）'
    ws.cell(row=1,column=1).alignment = alignment
    ws.cell(row=1,column=2).alignment = alignment
    ws.cell(row=1,column=3).alignment = alignment
    format(ws,4,18.0)
    for i in range(0,len(vola_range250)):
        ws.cell(row = i+1+1,column = 1).font = openpyxl.styles.Font(color=openpyxl.styles.colors.BLUE)
        ws.cell(row = i+1+1,column = 1).value = vola_range250[i][0]
        ws.cell(row = i+1+1, column = 2).value = vola_range250[i][1][0]
        ws.cell(row = i+1+1, column = 3).value = vola_range250[i][1][1]
    wb.save(today_name)
    # if(ALL == False):
    #     tkMessageBox.showinfo(title=u'写入成功', message=u'策略十四已经成功，请点击打开进行浏览')

# 15 #融资余额今10MA>前10MA,同时10MA>30MA


def strategy15():
    mrg = P.process15()
    wb = open_workbook(today_name)
    change_sheet(wb,u'股票代码')
    ws = open_sheet(wb,u'股票代码')
    ws.cell(row = 1,column = position15).value = u'融资余额'
    ws.cell(row = 1,column = position15).alignment = alignment
    for i in range(len(mrg)):
        ws.cell(row = i+1+1,column = position15).value = mrg[i][0]
        ws.cell(row = i+1+1,column = position15).font = openpyxl.styles.Font(color='00FF00FF')
    ws = open_sheet(wb,u'融资余额')
    ws.cell(row = 1, column = 1).value = u'符合条件代码'
    ws.cell(row = 1, column = 2).value = u'前一天融资余额\n'+ mrg[0][2][29].isoformat()[0:10]
    ws.cell(row = 1, column = 3).value = u'前一天融资余额10MA\n'+ mrg[0][2][29].isoformat()[0:10]
    ws.cell(row = 1, column = 4).value = u'前两天融资余额10MA\n'+ mrg[0][2][28].isoformat()[0:10]
    ws.cell(row = 1, column = 5).value = u'前一天融资余额30MA\n'+ mrg[0][2][28].isoformat()[0:10]
    ws.cell(row=1,column=1).alignment = alignment
    ws.cell(row=1,column=2).alignment = alignment
    ws.cell(row=1,column=3).alignment = alignment
    ws.cell(row=1,column=4).alignment = alignment
    ws.cell(row=1,column=5).alignment = alignment
    format(ws,5,17.0)
    for i in range(0,len(mrg)):
        ws.cell(row = i+1+1,column = 1).font = openpyxl.styles.Font(color=openpyxl.styles.colors.BLUE)
        ws.cell(row = i+1+1, column = 2).number_format = '#,##0.00_);[Green]-#,##0.00'
        ws.cell(row = i+1+1, column = 3).number_format = '#,##0.00_);[Green]-#,##0.00'
        ws.cell(row = i+1+1, column = 4).number_format = '#,##0.00_);[Green]-#,##0.00'
        ws.cell(row = i+1+1, column = 5).number_format = '#,##0.00_);[Green]-#,##0.00'
        ws.cell(row = i+1+1, column = 1).value = mrg[i][0]
        ws.cell(row = i+1+1, column = 2).value = mrg[i][1][0]
        ws.cell(row = i+1+1, column = 3).value = mrg[i][1][1]
        ws.cell(row = i+1+1, column = 4).value = mrg[i][1][2]
        ws.cell(row = i+1+1, column = 5).value = mrg[i][1][3]
    wb.save(today_name)
    # if(ALL == False):
    #     tkMessageBox.showinfo(title=u'写入成功', message=u'策略十五已经成功，请点击打开进行浏览')

# 16 #户均持股前三季中有一次增加


def strategy16():
    holder_avgnum = P.process16()
    wb = open_workbook(today_name)
    change_sheet(wb,u'股票代码')
    ws = open_sheet(wb,u'股票代码')
    ws.cell(row = 1,column = position16).value = u'户均持股'
    ws.cell(row = 1,column = position16).alignment = alignment
    for i in range(len(holder_avgnum)):
        ws.cell(row = i+1+1,column = position16).value = holder_avgnum[i][0]
        ws.cell(row = i+1+1,column = position16).font = openpyxl.styles.Font(color='00FF00FF')
    ws = open_sheet(wb,u'户均持股')
    ws.cell(row = 1, column = 1).value = u'符合条件代码'
    ws.cell(row = 1, column = 2).value = u'上三季户均持股'+'\n'+ holder_avgnum[0][2][0].isoformat()[0:10]
    ws.cell(row = 1, column = 3).value = u'上两季户均持股'+'\n'+ holder_avgnum[0][2][1].isoformat()[0:10]
    ws.cell(row = 1, column = 4).value = u'上一季户均持股'+'\n'+ holder_avgnum[0][2][2].isoformat()[0:10]
    ws.cell(row=1,column=1).alignment = alignment
    ws.cell(row=1,column=2).alignment = alignment
    ws.cell(row=1,column=3).alignment = alignment
    ws.cell(row=1,column=4).alignment = alignment
    format(ws,5,16.0)
    for i in range(0,len(holder_avgnum)):
        ws.cell(row = i+1+1,column = 1).font = openpyxl.styles.Font(color=openpyxl.styles.colors.BLUE)
        ws.cell(row = i+1+1, column = 2).number_format = '#,##0.00_);[Green]-#,##0.00'
        ws.cell(row = i+1+1, column = 3).number_format = '#,##0.00_);[Green]-#,##0.00'
        ws.cell(row = i+1+1, column = 4).number_format = '#,##0.00_);[Green]-#,##0.00'
        ws.cell(row = i+1+1, column = 1).value = holder_avgnum[i][0]
        ws.cell(row = i+1+1, column = 2).value = holder_avgnum[i][1][0]
        ws.cell(row = i+1+1, column = 3).value = holder_avgnum[i][1][1]
        ws.cell(row = i+1+1, column = 4).value = holder_avgnum[i][1][2]
    wb.save(today_name)
    # if(ALL == False):
    #     tkMessageBox.showinfo(title=u'写入成功', message=u'策略十六已经成功，请点击打开进行浏览')

# 17 #股东户数两季连续减少


def strategy17():
    holder_num = P.process17()
    wb = open_workbook(today_name)
    change_sheet(wb,u'股票代码')
    ws = open_sheet(wb,u'股票代码')
    ws.cell(row = 1,column = position17).value = u'股东户数'
    ws.cell(row = 1,column = position17).alignment = alignment
    for i in range(len(holder_num)):
        ws.cell(row = i+1+1,column = position17).value = holder_num[i][0]
        ws.cell(row = i+1+1,column = position17).font = openpyxl.styles.Font(color='00FF00FF')
    ws = open_sheet(wb,u'股东户数')
    ws.cell(row = 1, column = 1).value = u'符合条件代码'
    ws.cell(row = 1, column = 2).value = u'上三季股东户数'+'\n'+ holder_num[0][2][0].isoformat()[0:10]
    ws.cell(row = 1, column = 3).value = u'上两季股东户数'+'\n'+ holder_num[0][2][1].isoformat()[0:10]
    ws.cell(row = 1, column = 4).value = u'上一季股东户数'+'\n'+ holder_num[0][2][2].isoformat()[0:10]
    ws.cell(row=1,column=1).alignment = alignment
    ws.cell(row=1,column=2).alignment = alignment
    ws.cell(row=1,column=3).alignment = alignment
    ws.cell(row=1,column=4).alignment = alignment
    format(ws,5,16.0)
    for i in range(0,len(holder_num)):
        ws.cell(row = i+1+1,column = 1).font = openpyxl.styles.Font(color=openpyxl.styles.colors.BLUE)
        ws.cell(row = i+1+1, column = 2).number_format = '#,##_);[Green]-#,##'
        ws.cell(row = i+1+1, column = 3).number_format = '#,##_);[Green]-#,##'
        ws.cell(row = i+1+1, column = 4).number_format = '#,##_);[Green]-#,##'
        ws.cell(row = i+1+1, column = 1).value = holder_num[i][0]
        ws.cell(row = i+1+1, column = 2).value = holder_num[i][1][0]
        ws.cell(row = i+1+1, column = 3).value = holder_num[i][1][1]
        ws.cell(row = i+1+1, column = 4).value = holder_num[i][1][2]
    wb.save(today_name)
    # if(ALL == False):
    #     tkMessageBox.showinfo(title=u'写入成功', message=u'策略十七已经成功，请点击打开进行浏览')

# 18 #十大股东变化超过3个


def strategy18():
    change = P.process18()
    wb = open_workbook(today_name)
    change_sheet(wb,u'股票代码')
    ws = open_sheet(wb,u'股票代码')
    ws.cell(row = 1,column = position18).value = u'十大股东变化数'
    ws.cell(row = 1,column = position18).alignment = alignment
    for i in range(len(change)):
        ws.cell(row = i+1+1,column = position18).value = change[i][0]
        ws.cell(row = i+1+1,column = position18).font = openpyxl.styles.Font(color='00FF00FF')
    ws = open_sheet(wb,u'十大股东变化数')
    ws.cell(row = 1, column = 1).value = u'符合条件代码'
    ws.cell(row = 1, column = 2).value = u'上季与上两季变化数量'
    ws.cell(row = 1, column = 3).value = u'上两季名单'+'\n'+ change[0][2][1].isoformat()[0:10]
    ws.cell(row = 1, column = 4).value = u'上一季名单'+'\n'+ change[0][2][2].isoformat()[0:10]
    ws.cell(row=1,column=1).alignment = alignment
    ws.cell(row=1,column=2).alignment = alignment
    ws.cell(row=1,column=3).alignment = alignment
    ws.cell(row=1,column=4).alignment = alignment
    format(ws,5,16.0)
    for i in range(0,len(change)):
        ws.cell(row = i+1+1,column = 1).font = openpyxl.styles.Font(color=openpyxl.styles.colors.BLUE)
        ws.cell(row = i+1+1, column = 1).value = change[i][0]
        ws.cell(row = i+1+1, column = 2).value = change[i][3]
        ws.cell(row = i+1+1, column = 3).value = change[i][1][0]
        ws.cell(row = i+1+1, column = 4).value = change[i][1][1]
    wb.save(today_name)
    # if(ALL == False):
    #     tkMessageBox.showinfo(title=u'写入成功', message=u'策略十八已经成功，请点击打开进行浏览')

# 19 #十大股东持股比例增加


def strategy19():
    holder_top10, all = P.process19()
    wb = open_workbook(today_name)

    if len(holder_top10):
        last_three = holder_top10[0][2][0].isoformat()[0:10]
        last_two = holder_top10[0][2][1].isoformat()[0:10]
        last_one = holder_top10[0][2][2].isoformat()[0:10]
    else:
        last_three = str(' ')
        last_two = str(' ')
        last_one = str(' ')
    change_sheet(wb,u'股票代码')
    ws = open_sheet(wb,u'股票代码')
    ws.cell(row = 1,column = position19).value = u'十大股东持股比例'
    ws.cell(row = 1,column = position19).alignment = alignment
    for i in range(len(holder_top10)):
        ws.cell(row = i+1+1,column = position19).value = holder_top10[i][0]
        ws.cell(row = i+1+1,column = position19).font = openpyxl.styles.Font(color='00FF00FF')
    ws = open_sheet(wb,u'十大股东持股比例')
    ws.cell(row = 1, column = 1).value = u'符合条件代码'
    ws.cell(row = 1, column = 2).value = u'上三季十大股东持股比例'+'\n'+ last_three
    ws.cell(row = 1, column = 3).value = u'上两季十大股东持股比例'+'\n'+ last_two
    ws.cell(row = 1, column = 4).value = u'上一季十大股东持股比例'+'\n'+ last_one
    ws.cell(row = 1, column = 5).value = u'全部股票'
    ws.cell(row = 1, column = 6).value = u'上两季十大股东持股比例\n（全部）'+'\n'+ last_one
    ws.cell(row = 1, column = 7).value = u'上一季十大股东持股比例\n（全部）'+'\n'+ last_one
    ws.cell(row = 1, column = 8).value = u'排序'+'\n'+ last_one
    ws.cell(row=1,column=1).alignment = alignment
    ws.cell(row=1,column=2).alignment = alignment
    ws.cell(row=1,column=3).alignment = alignment
    ws.cell(row=1,column=4).alignment = alignment
    ws.cell(row=1,column=5).alignment = alignment
    ws.cell(row=1,column=6).alignment = alignment
    ws.cell(row=1,column=7).alignment = alignment
    ws.cell(row=1,column=8).alignment = alignment
    format(ws,8,26.0)
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['H'].width = 20
    for i in range(0,len(holder_top10)):
        ws.cell(row = i+1+1,column = 1).font = openpyxl.styles.Font(color=openpyxl.styles.colors.BLUE)
        ws.cell(row = i+1+1, column = 2).number_format = '#,##0.00_)\%;[Green]-#,##0.00\%'
        ws.cell(row = i+1+1, column = 3).number_format = '#,##0.00_)\%;[Green]-#,##0.00\%'
        ws.cell(row = i+1+1, column = 4).number_format = '#,##0.00_)\%;[Green]-#,##0.00\%'
        ws.cell(row = i+1+1, column = 1).value = holder_top10[i][0]
        ws.cell(row = i+1+1, column = 2).value = holder_top10[i][1][0]
        ws.cell(row = i+1+1, column = 3).value = holder_top10[i][1][1]
        ws.cell(row = i+1+1, column = 4).value = holder_top10[i][1][2]
    for i in range(len(all)):
        ws.cell(row=i + 1 + 1, column=5).font = openpyxl.styles.Font(color=openpyxl.styles.colors.BLUE)
        ws.cell(row=i + 1 + 1, column=6).number_format = '#,##0.00_)\%;[Green]-#,##0.00\%'
        ws.cell(row=i + 1 + 1, column=7).number_format = '#,##0.00_)\%;[Green]-#,##0.00\%'
        ws.cell(row=i + 1 + 1, column=8).number_format = '#,##0.00_)%;[Green]-#,##0.00%'
        ws.cell(row=i + 1 + 1, column=5).value = all[i][0]
        ws.cell(row=i + 1 + 1, column=6).value = all[i][1] - 1
        ws.cell(row=i + 1 + 1, column=7).value = all[i][2] - 1
        ws.cell(row=i + 1 + 1, column=8).value = all[i][3] - 1

    wb.save(today_name)
    # if(ALL == False):
    #     tkMessageBox.showinfo(title=u'写入成功', message=u'策略十九已经成功，请点击打开进行浏览')
# 20、21、22 #券商、基金、机构持股数量排序


def strategy202122():
    trader,fund,institude = P.process202122()
    wb = open_workbook(today_name)
    change_sheet(wb,u'股票代码')
    ws = open_sheet(wb,u'股票代码')
    ws.cell(row = 1,column = position20).value = u'券商'
    ws.cell(row = 1,column = position20).alignment = alignment
    for i in range(len(trader)):
        ws.cell(row = i+1+1,column = position20).value = trader[i][0]
        ws.cell(row = i+1+1,column = position20).font = openpyxl.styles.Font(color='00FF00FF')
    ws.cell(row = 1,column = position21).value = u'基金'
    ws.cell(row = 1,column = position21).alignment = alignment
    for i in range(len(fund)):
        ws.cell(row = i+1+1,column = position21).value = fund[i][0]
        ws.cell(row = i+1+1,column = position21).font = openpyxl.styles.Font(color='00FF00FF')
    ws.cell(row = 1,column = position22).value = u'机构'
    ws.cell(row = 1,column = position22).alignment = alignment
    for i in range(len(institude)):
        ws.cell(row = i+1+1,column = position22).value = institude[i][0]
        ws.cell(row = i+1+1,column = position22).font = openpyxl.styles.Font(color='00FF00FF')
    ws = open_sheet(wb,u'券商、基金、机构')
    ws.cell(row = 1, column = 1).value = u'代码\n（由高到低排序）'
    ws.cell(row = 1, column = 2).value = u'券商持股数'+'\n'+ trader[0][2][0].isoformat()[0:10]
    ws.cell(row = 1, column = 3).value = u'代码\n（由高到低排序）'
    ws.cell(row = 1, column = 4).value = u'基金持股数'+'\n'+ fund[0][2][0].isoformat()[0:10]
    ws.cell(row = 1, column = 5).value = u'代码\n（由高到低排序）'
    ws.cell(row = 1, column = 6).value = u'机构持股数'+'\n'+ institude[0][2][0].isoformat()[0:10]
    ws.cell(row=1,column=1).alignment = alignment
    ws.cell(row=1,column=2).alignment = alignment
    ws.cell(row=1,column=3).alignment = alignment
    ws.cell(row=1,column=4).alignment = alignment
    ws.cell(row=1,column=5).alignment = alignment
    ws.cell(row=1,column=6).alignment = alignment
    format(ws,6,18.0)
    ws.column_dimensions['A'].width = 18
    for i in range(0,len(trader)):
        ws.cell(row = i+1+1,column = 1).font = openpyxl.styles.Font(color=openpyxl.styles.colors.BLUE)
        ws.cell(row = i+1+1, column = 1).value = trader[i][0]
        ws.cell(row = i+1+1, column = 2).value = trader[i][1]
        ws.cell(row = i+1+1, column = 2).number_format = '#,##_);[Green]-#,##'
    for i in range(0,len(fund)):
        ws.cell(row = i+1+1,column = 3).font = openpyxl.styles.Font(color=openpyxl.styles.colors.BLUE)
        ws.cell(row = i+1+1, column = 3).value = fund[i][0]
        ws.cell(row = i+1+1, column = 4).value = fund[i][1]
        ws.cell(row = i+1+1, column = 4).number_format = '#,##_);[Green]-#,##'
    for i in range(0,len(institude)):
        ws.cell(row = i+1+1,column = 5).font = openpyxl.styles.Font(color=openpyxl.styles.colors.BLUE)
        ws.cell(row = i+1+1, column = 5).value = institude[i][0]
        ws.cell(row = i+1+1, column = 6).value = institude[i][1]
        ws.cell(row = i+1+1, column = 6).number_format = '#,##_);[Green]-#,##'
    wb.save(today_name)
    # if(ALL == False):
    #     tkMessageBox.showinfo(title=u'写入成功', message=u'策略二十已经成功，请点击打开进行浏览')

# 23 #换手率5MA大于前一天5MA，同时10MA大于前一天10MA


def strategy23():
    turn_num = P.process23()
    wb = open_workbook(today_name)
    change_sheet(wb,u'股票代码')
    ws = open_sheet(wb,u'股票代码')
    ws.cell(row = 1,column = position23).value = u'换手率'
    ws.cell(row = 1,column = position23).alignment = alignment
    for i in range(len(turn_num)):
        ws.cell(row = i+1+1,column = position23).value = turn_num[i][0]
        ws.cell(row = i+1+1,column = position23).font = openpyxl.styles.Font(color='00FF00FF')
    ws = open_sheet(wb,u'换手率')
    ws.cell(row = 1, column = 1).value = u'符合条件代码'
    ws.cell(row = 1, column = 2).value = u'今日换手率\n' + turn_num[0][2][10].isoformat()[0:10]
    ws.cell(row = 1, column = 3).value = u'前一天换手率5MA\n' + turn_num[0][2][9].isoformat()[0:10]
    ws.cell(row = 1, column = 4).value = u'今日换手率5MA\n' + turn_num[0][2][10].isoformat()[0:10]
    ws.cell(row = 1, column = 5).value = u'前一天换手率10MA\n' + turn_num[0][2][9].isoformat()[0:10]
    ws.cell(row = 1, column = 6).value = u'今日换手率10MA\n' + turn_num[0][2][10].isoformat()[0:10]
    ws.cell(row=1,column=1).alignment = alignment
    ws.cell(row=1,column=2).alignment = alignment
    ws.cell(row=1,column=3).alignment = alignment
    ws.cell(row=1,column=4).alignment = alignment
    ws.cell(row=1,column=5).alignment = alignment
    ws.cell(row=1,column=6).alignment = alignment
    format(ws,6,15.0)
    for i in range(0,len(turn_num)):
        ws.cell(row = i+1+1,column = 1).font = openpyxl.styles.Font(color=openpyxl.styles.colors.BLUE)
        ws.cell(row = i+1+1, column = 2).number_format = '#,##0.00_)\%;[Green]-#,##0.00\%'
        ws.cell(row = i+1+1, column = 3).number_format = '#,##0.00_)\%;[Green]-#,##0.00\%'
        ws.cell(row = i+1+1, column = 4).number_format = '#,##0.00_)\%;[Green]-#,##0.00\%'
        ws.cell(row = i+1+1, column = 5).number_format = '#,##0.00_)\%;[Green]-#,##0.00\%'
        ws.cell(row = i+1+1, column = 6).number_format = '#,##0.00_)\%;[Green]-#,##0.00\%'
        ws.cell(row = i+1+1,column = 1).value = turn_num[i][0]
        ws.cell(row = i+1+1, column = 2).value = turn_num[i][1][0]
        ws.cell(row = i+1+1, column = 3).value = turn_num[i][1][1]
        ws.cell(row = i+1+1, column = 4).value = turn_num[i][1][2]
        ws.cell(row = i+1+1, column = 5).value = turn_num[i][1][3]
        ws.cell(row = i+1+1, column = 6).value = turn_num[i][1][4]
    wb.save(today_name)
    # if(ALL == False):
        # tkMessageBox.showinfo(title=u'写入成功', message=u'策略二十一已经成功，请点击打开进行浏览')
# 24 #涨幅日、周、月都超过大盘


def strategy24():
    increase = P.process24()
    if len(increase):
        last_one = increase[0][2][19].isoformat()[0:10]
        last_five = increase[0][2][15].isoformat()[0:10]
        last_twenty = increase[0][2][0].isoformat()[0:10]
    else:
        last_one = str(' ')
        last_five = str(' ')
        last_twenty = str(' ')
    wb = open_workbook(today_name)
    change_sheet(wb,u'股票代码')
    ws = open_sheet(wb,u'股票代码')
    ws.cell(row = 1,column = position24).value = u'涨幅'
    ws.cell(row = 1,column = position24).alignment = alignment
    for i in range(len(increase)):
        ws.cell(row=i + 1 + 1, column=position24).value = increase[i][0]
        ws.cell(row=i + 1 + 1, column=position24).font = openpyxl.styles.Font(color=openpyxl.styles.colors.RED)
    ws = open_sheet(wb, u'涨幅')
    ws.cell(row = 1, column = 1).value = u'符合条件代码'
    ws.cell(row = 1, column = 2).value = u'日涨幅\n（前一天）\n' + last_one
    ws.cell(row = 1, column = 3).value = u'周涨幅\n（前五天）\n' + last_five
    ws.cell(row = 1, column = 4).value = u'月涨幅\n（二十天）\n' + last_twenty
    ws.cell(row = 1, column = 5).value = u'大盘日涨幅\n（前一天）\n' + last_one
    ws.cell(row = 1, column = 6).value = u'大盘周涨幅\n（前五天）\n' + last_five
    ws.cell(row = 1, column = 7).value = u'大盘月涨幅\n（二十天）\n' + last_twenty
    ws.cell(row=1,column=1).alignment = alignment
    ws.cell(row=1,column=2).alignment = alignment
    ws.cell(row=1,column=3).alignment = alignment
    ws.cell(row=1,column=4).alignment = alignment
    ws.cell(row=1,column=5).alignment = alignment
    ws.cell(row=1,column=6).alignment = alignment
    ws.cell(row=1,column=7).alignment = alignment
    format(ws, 7, 15.0)
    for i in range(0,len(increase)):
        ws.cell(row = i+1+1,column = 1).font = openpyxl.styles.Font(color=openpyxl.styles.colors.BLUE)
        ws.cell(row = i+1+1,column = 2).font = openpyxl.styles.Font(color=openpyxl.styles.colors.RED)
        ws.cell(row = i+1+1,column = 3).font = openpyxl.styles.Font(color=openpyxl.styles.colors.RED)
        ws.cell(row = i+1+1,column = 4).font = openpyxl.styles.Font(color=openpyxl.styles.colors.RED)
        ws.cell(row = i+1+1, column = 2).number_format = '#,##0.00%_);[Green]-#,##0.00%'
        ws.cell(row = i+1+1, column = 3).number_format = '#,##0.00%_);[Green]-#,##0.00%'
        ws.cell(row = i+1+1, column = 4).number_format = '#,##0.00%_);[Green]-#,##0.00%'
        ws.cell(row = i+1+1, column = 5).number_format = '#,##0.00%_);[Green]-#,##0.00%'
        ws.cell(row = i+1+1, column = 6).number_format = '#,##0.00%_);[Green]-#,##0.00%'
        ws.cell(row = i+1+1, column = 7).number_format = '#,##0.00%_);[Green]-#,##0.00%'
        ws.cell(row = i+1+1,column = 1).value = increase[i][0]
        ws.cell(row = i+1+1, column = 2).value = increase[i][1][0]
        ws.cell(row = i+1+1, column = 3).value = increase[i][1][1]
        ws.cell(row = i+1+1, column = 4).value = increase[i][1][2]
        ws.cell(row = i+1+1, column = 5).value = increase[i][1][3]
        ws.cell(row = i+1+1, column = 6).value = increase[i][1][4]
        ws.cell(row = i+1+1, column = 7).value = increase[i][1][4]
    wb.save(today_name)
# 25 #袁氏选股


def strategy25():
    yuan1,yuan2,all = P.process25()
    if len(yuan1):
        day_today = yuan1[0][2][252].isoformat()[0:10]
        day_pre_5 = yuan1[0][2][247].isoformat()[0:10]
        day_pre_60 = yuan1[0][2][192].isoformat()[0:10]
    else:
        day_today = str(' ')
        day_pre_5 = str(' ')
        day_pre_60 = str(' ')
    if len(yuan2):
        day_today = yuan2[0][2][252].isoformat()[0:10]
        day_pre_5 = yuan2[0][2][247].isoformat()[0:10]
        dau_pre_252 = yuan2[0][2][0].isoformat()[0:10]
    else:
        day_today = str(' ')
        day_pre_5 = str(' ')
        dau_pre_252 = str(' ')
    if len(all):
        day_today = yuan1[0][2][252].isoformat()[0:10]
        day_pre_5 = yuan1[0][2][247].isoformat()[0:10]
        day_pre_60 = yuan1[0][2][192].isoformat()[0:10]
        dau_pre_252 = yuan1[0][2][0].isoformat()[0:10]
    else:
        day_today = str(' ')
        day_pre_5 = str(' ')
        day_pre_60 = str(' ')
        dau_pre_252 = str(' ')
    wb = open_workbook(today_name)
    change_sheet(wb, u'股票代码')
    ws = open_sheet(wb, u'股票代码')
    ws.cell(row = 1,column = position25).value = u'袁氏选股\n(积极版)'
    ws.cell(row = 1,column = position25).alignment = alignment
    ws.cell(row = 1,column = position26).value = u'袁氏选股\n(年版)'
    ws.cell(row = 1,column = position26).alignment = alignment
    for i in range(len(yuan1)):
        ws.cell(row=i + 1 + 1, column=position25).value = yuan1[i][0]
        ws.cell(row=i + 1 + 1, column=position25).font = openpyxl.styles.Font(color=openpyxl.styles.colors.RED)
    for i in range(len(yuan2)):
        ws.cell(row=i + 1 + 1, column=position26).value = yuan2[i][0]
        ws.cell(row=i + 1 + 1, column=position26).font = openpyxl.styles.Font(color=openpyxl.styles.colors.RED)
    ws = open_sheet(wb, u'袁氏选股')
    ws.cell(row=1, column=1).value = u'符合条件代码（积极版）'
    ws.cell(row = 1, column = 2).value = u'收盘价\n（当日）\n' + day_today
    ws.cell(row = 1, column = 3).value = u'收盘价\n（前五天）\n' + day_pre_5
    ws.cell(row = 1, column = 4).value = u'收盘价\n（前六十天）\n' + day_pre_60
    ws.cell(row=1, column=5).value = u'符合条件代码（年版）'
    ws.cell(row = 1, column = 6).value = u'收盘价\n（当日）\n' + day_today
    ws.cell(row = 1, column = 7).value = u'收盘价\n（前五天）\n' + day_pre_5
    ws.cell(row = 1, column = 8).value = u'收盘价\n（前252天）\n' + dau_pre_252
    ws.cell(row = 1, column = 9).value = u'全部代码'
    ws.cell(row = 1, column = 10).value = u'收盘价\n（当日）\n' + day_today
    ws.cell(row = 1, column = 11).value = u'收盘价\n（前五天）\n' + day_pre_5
    ws.cell(row = 1, column = 12).value = u'收盘价\n（前六十天）\n' + day_pre_60
    ws.cell(row = 1, column = 13).value = u'收盘价\n（前252天）\n' + dau_pre_252
    ws.cell(row=1,column=1).alignment = alignment
    ws.cell(row=1,column=2).alignment = alignment
    ws.cell(row=1,column=3).alignment = alignment
    ws.cell(row=1,column=4).alignment = alignment
    ws.cell(row=1,column=5).alignment = alignment
    ws.cell(row=1,column=6).alignment = alignment
    ws.cell(row=1,column=7).alignment = alignment
    ws.cell(row=1,column=8).alignment = alignment
    ws.cell(row=1,column=9).alignment = alignment
    ws.cell(row=1,column=10).alignment = alignment
    ws.cell(row=1,column=11).alignment = alignment
    ws.cell(row=1,column=12).alignment = alignment
    ws.cell(row=1,column=13).alignment = alignment
    ws.column_dimensions['A'].width = 15.0
    format(ws, 13, 15.0)
    for i in range(0, len(yuan1)):
        ws.cell(row=1, column=1).alignment = alignment
        ws.cell(row = i+1+1,column = 1).font = openpyxl.styles.Font(color=openpyxl.styles.colors.BLUE)
        ws.cell(row = i+1+1,column = 2).font = openpyxl.styles.Font(color=openpyxl.styles.colors.RED)
        ws.cell(row = i+1+1,column = 3).font = openpyxl.styles.Font(color=openpyxl.styles.colors.RED)
        ws.cell(row = i+1+1,column = 4).font = openpyxl.styles.Font(color=openpyxl.styles.colors.RED)
        ws.cell(row = i+1+1,column = 1).value = yuan1[i][0]
        ws.cell(row = i+1+1, column = 2).value = yuan1[i][1][0]
        ws.cell(row = i+1+1, column = 3).value = yuan1[i][1][1]
        ws.cell(row = i+1+1, column = 4).value = yuan1[i][1][2]
    for i in range(0, len(yuan2)):
        ws.cell(row=1, column=5).alignment = alignment
        ws.cell(row = i+1+1,column = 5).font = openpyxl.styles.Font(color=openpyxl.styles.colors.BLUE)
        ws.cell(row = i+1+1,column = 6).font = openpyxl.styles.Font(color=openpyxl.styles.colors.RED)
        ws.cell(row = i+1+1,column = 7).font = openpyxl.styles.Font(color=openpyxl.styles.colors.RED)
        ws.cell(row = i+1+1,column = 8).font = openpyxl.styles.Font(color=openpyxl.styles.colors.RED)
        ws.cell(row = i+1+1, column = 5).value = yuan2[i][0]
        ws.cell(row = i+1+1, column = 6).value = yuan2[i][1][0]
        ws.cell(row = i+1+1, column = 7).value = yuan2[i][1][1]
        ws.cell(row = i+1+1, column = 8).value = yuan2[i][1][2]
    for i in range(0,len(all)):
        ws.cell(row=1, column=9).alignment = alignment
        ws.cell(row=i + 1 + 1, column=9).font = openpyxl.styles.Font(color=openpyxl.styles.colors.BLUE)
        ws.cell(row=i + 1 + 1, column=10).font = openpyxl.styles.Font(color=openpyxl.styles.colors.RED)
        ws.cell(row=i + 1 + 1, column=11).font = openpyxl.styles.Font(color=openpyxl.styles.colors.RED)
        ws.cell(row=i + 1 + 1, column=12).font = openpyxl.styles.Font(color=openpyxl.styles.colors.RED)
        ws.cell(row=i + 1 + 1, column=13).font = openpyxl.styles.Font(color=openpyxl.styles.colors.RED)
        ws.cell(row = i+1+1, column = 9).value = all[i][0]
        ws.cell(row = i+1+1, column = 10).value = all[i][1]
        ws.cell(row = i+1+1, column = 11).value = all[i][2]
        ws.cell(row = i+1+1, column = 12).value = all[i][3]
        ws.cell(row = i+1+1, column = 13).value = all[i][4]

    wb.save(today_name)
# 27 # 股本，总股本和流通股本


def strategy27():
    total_share,float_share = P.process27()
    wb = open_workbook(today_name)
    change_sheet(wb, u'股票代码')
    ws = open_sheet(wb, u'股票代码')
    ws.cell(row=1, column=position27).value = u'总股本'
    ws.cell(row=1, column=position27).alignment = alignment
    for i in range(len(total_share)):
        ws.cell(row=i + 1 + 1, column=position27).value = total_share[i][0]
        ws.cell(row=i + 1 + 1, column=position27).font = openpyxl.styles.Font(color='00FF00FF')
    ws.cell(row=1, column=position28).value = u'流通股本'
    ws.cell(row=1, column=position28).alignment = alignment
    for i in range(len(float_share)):
        ws.cell(row=i + 1 + 1, column=position28).value = float_share[i][0]
        ws.cell(row=i + 1 + 1, column=position28).font = openpyxl.styles.Font(color='00FF00FF')
    ws = open_sheet(wb, u'股本资料')
    ws.cell(row=1, column=1).value = u'代码\n（由高到低排序）'
    ws.cell(row=1, column=2).value = u'总股本' + '\n' + total_share[0][2][0].isoformat()[0:10]
    ws.cell(row=1, column=3).value = u'代码\n（由高到低排序）'
    ws.cell(row=1, column=4).value = u'流通股本' + '\n' + float_share[0][2][0].isoformat()[0:10]
    ws.cell(row=1, column=1).alignment = alignment
    ws.cell(row=1, column=2).alignment = alignment
    ws.cell(row=1, column=3).alignment = alignment
    ws.cell(row=1, column=4).alignment = alignment
    format(ws, 6, 18.0)
    ws.column_dimensions['A'].width = 18
    for i in range(0, len(total_share)):
        ws.cell(row=i + 1 + 1, column=1).font = openpyxl.styles.Font(color=openpyxl.styles.colors.BLUE)
        ws.cell(row=i + 1 + 1, column=1).value = total_share[i][0]
        ws.cell(row=i + 1 + 1, column=2).value = total_share[i][1]
        ws.cell(row=i + 1 + 1, column=2).number_format = '#,##_);[Green]-#,##'
    for i in range(0, len(float_share)):
        ws.cell(row=i + 1 + 1, column=3).font = openpyxl.styles.Font(color=openpyxl.styles.colors.BLUE)
        ws.cell(row=i + 1 + 1, column=3).value = float_share[i][0]
        ws.cell(row=i + 1 + 1, column=4).value = float_share[i][1]
        ws.cell(row=i + 1 + 1, column=4).number_format = '#,##_);[Green]-#,##'
    wb.save(today_name)
# 29 # 公司净值（每股净资产BPS）
def strategy29():
    bps = P.process29()
    if len(bps):
        day_today = bps[0][2][0].isoformat()[0:10]
    else:
        day_today = str(' ')
    wb = open_workbook(today_name)
    change_sheet(wb, u'股票代码')
    ws = open_sheet(wb, u'股票代码')
    ws.cell(row=1, column=position29).value = u'净值（每股净资产BPS）'
    ws.cell(row=1, column=position29).alignment = alignment
    for i in range(len(bps)):
        ws.cell(row=i + 1 + 1, column=position29).value = bps[i][0]
        ws.cell(row=i + 1 + 1, column=position29).font = openpyxl.styles.Font(color='32CD32')
    ws = open_sheet(wb, u'股本资料')
    ws.cell(row=1, column=1).value = u'代码\n（由高到低排序）'
    ws.cell(row=1, column=2).value = u'每股净资产(元)' + '\n' + day_today
    ws.cell(row=1, column=1).alignment = alignment
    ws.cell(row=1, column=2).alignment = alignment
    format(ws, 6, 18.0)
    ws.column_dimensions['A'].width = 18
    for i in range(0, len(bps)):
        ws.cell(row=i + 1 + 1, column=1).font = openpyxl.styles.Font(color=openpyxl.styles.colors.BLUE)
        ws.cell(row=i + 1 + 1, column=1).value = bps[i][0]
        ws.cell(row=i + 1 + 1, column=2).value = bps[i][1]
        ws.cell(row=i + 1 + 1, column=2).number_format = '#,##0.00_);[Green]-#,##0.00'
    wb.save(today_name)
# 30 # 市盈率PE
def strategy30():
    pe = P.process30()
    if len(pe):
        day_today = pe[0][2][0].isoformat()[0:10]
    else:
        day_today = str(' ')
    wb = open_workbook(today_name)
    change_sheet(wb, u'股票代码')
    ws = open_sheet(wb, u'股票代码')
    ws.cell(row=1, column=position30).value = u'市盈率（PE）'
    ws.cell(row=1, column=position30).alignment = alignment
    for i in range(len(pe)):
        ws.cell(row=i + 1 + 1, column=position30).value = pe[i][0]
        ws.cell(row=i + 1 + 1, column=position30).font = openpyxl.styles.Font(color='32CD32')
    ws = open_sheet(wb, u'市盈率')
    ws.cell(row=1, column=1).value = u'代码\n（由高到低排序）'
    ws.cell(row=1, column=2).value = u'市盈率（倍）' + '\n' + day_today
    ws.cell(row=1, column=1).alignment = alignment
    ws.cell(row=1, column=2).alignment = alignment
    format(ws, 6, 18.0)
    ws.column_dimensions['A'].width = 18
    for i in range(0, len(pe)):
        ws.cell(row=i + 1 + 1, column=1).font = openpyxl.styles.Font(color=openpyxl.styles.colors.BLUE)
        ws.cell(row=i + 1 + 1, column=1).value = pe[i][0]
        ws.cell(row=i + 1 + 1, column=2).value = pe[i][1]
        ws.cell(row=i + 1 + 1, column=2).number_format = '#,##0.00_);[Green]-#,##0.00'
    wb.save(today_name)
# 31 # 每股分红送转
def strategy31():
    fenhong = P.process31()
    if len(fenhong):
        day_today = fenhong[0][2][0].isoformat()[0:10]
    else:
        day_today = str(' ')
    wb = open_workbook(today_name)
    change_sheet(wb, u'股票代码')
    ws = open_sheet(wb, u'股票代码')
    ws.cell(row=1, column=position31).value = u'每股分红送转'
    ws.cell(row=1, column=position31).alignment = alignment
    for i in range(len(fenhong)):
        ws.cell(row=i + 1 + 1, column=position31).value = fenhong[i][0]
        ws.cell(row=i + 1 + 1, column=position31).font = openpyxl.styles.Font(color='32CD32')
    ws = open_sheet(wb, u'分红转送')
    ws.cell(row=1, column=1).value = u'代码\n（由高到低排序）'
    ws.cell(row=1, column=2).value = u'每股分红送转(元)' + '\n' + day_today
    ws.cell(row=1, column=1).alignment = alignment
    ws.cell(row=1, column=2).alignment = alignment
    format(ws, 6, 18.0)
    ws.column_dimensions['A'].width = 18
    for i in range(0, len(fenhong)):
        ws.cell(row=i + 1 + 1, column=1).font = openpyxl.styles.Font(color=openpyxl.styles.colors.BLUE)
        ws.cell(row=i + 1 + 1, column=1).value = fenhong[i][0]
        ws.cell(row=i + 1 + 1, column=2).value = fenhong[i][1]
        ws.cell(row=i + 1 + 1, column=2).number_format = '#,##0.00_);[Green]-#,##0.00'
    wb.save(today_name)

# 期货策略 #
# 5 #波动率今10MA > 前10MA，同时 10 MA > 100MA


def future_strategy5():
    volatility = P.process5()
    wb = open_workbook(today_name_future)
    change_sheet(wb,u'期货代码')
    ws = open_sheet(wb,u'期货代码')
    ft = openpyxl.styles.Font(color=openpyxl.styles.colors.RED)
    ws.cell(row = 1,column = 1).value = u'波动率'
    ws.cell(row = 1,column = 1).alignment = alignment
    for i in range(len(volatility)):
        ws.cell(row = i+1+1,column = 1).value = volatility[i][0]
        ws.cell(row = i+1+1,column = 1).font = openpyxl.styles.Font(color=openpyxl.styles.colors.RED)
    ws = open_sheet(wb,u'波动率')
    ws.cell(row = 1, column = 1).value = u'符合条件代码'
    ws.cell(row = 1, column = 2).value = u'今日波动率\n(10天为周期)\n' + volatility[0][2][110].isoformat()[0:10]
    ws.cell(row = 1, column = 3).value = u'前一天10MA'
    ws.cell(row = 1, column = 4).value = u'今日10MA'
    ws.cell(row = 1, column = 5).value = u'100MA'
    ws.cell(row=1,column=1).alignment = alignment
    ws.cell(row=1,column=2).alignment = alignment
    ws.cell(row=1,column=3).alignment = alignment
    ws.cell(row=1,column=4).alignment = alignment
    ws.cell(row=1,column=5).alignment = alignment
    format(ws,5,15.0)
    for i in range(0,len(volatility)):
        ws.cell(row = i+1+1,column = 1).font = openpyxl.styles.Font(color=openpyxl.styles.colors.BLUE)
        ws.cell(row = i+1+1, column = 2).number_format = '0.00000%;[Green]-0.00000%'
        ws.cell(row = i+1+1, column = 3).number_format = '0.00000%;[Green]-0.00000%'
        ws.cell(row = i+1+1, column = 4).number_format = '0.00000%;[Green]-0.00000%'
        ws.cell(row = i+1+1, column = 5).number_format = '0.00000%;[Green]-0.00000%'
        ws.cell(row = i+1+1, column = 1).value = volatility[i][0]
        ws.cell(row = i+1+1, column = 2).value = volatility[i][1][0]
        ws.cell(row = i+1+1, column = 3).value = volatility[i][1][1]
        ws.cell(row = i+1+1, column = 4).value = volatility[i][1][2]
        ws.cell(row = i+1+1, column = 5).value = volatility[i][1][3]
    wb.save(today_name_future)
    # if(ALL == False):
    #     tkMessageBox.showinfo(title=u'写入成功', message=u'策略五已经成功，请点击打开进行浏览')

# 6 #价格今10MA > 前10MA，同时前10MA > 前前10MA


def future_strategy6():
    price = P.process6()
    wb = open_workbook(today_name_future)
    change_sheet(wb,u'期货代码')
    ws = open_sheet(wb,u'期货代码')
    ws.cell(row = 1,column = 2).value = u'价格MA'
    ws.cell(row = 1,column = 2).alignment = alignment
    for i in range(len(price)):
        ws.cell(row = i+1+1,column = 2).value = price[i][0]
        ws.cell(row = i+1+1,column = 2).font = openpyxl.styles.Font(color=openpyxl.styles.colors.RED)
    ws = open_sheet(wb,u'价格MA')
    ws.cell(row = 1, column = 1).value = u'符合条件代码'
    ws.cell(row = 1, column = 2).value = u'前两天10MA\n' + price[0][2][9].isoformat()[0:10]
    ws.cell(row = 1, column = 3).value = u'前一天10MA\n' + price[0][2][10].isoformat()[0:10]
    ws.cell(row = 1, column = 4).value = u'今日10MA\n' + price[0][2][11].isoformat()[0:10]
    ws.cell(row=1,column=1).alignment = alignment
    ws.cell(row=1,column=2).alignment = alignment
    ws.cell(row=1,column=3).alignment = alignment
    ws.cell(row=1,column=4).alignment = alignment
    format(ws,4,15.0)
    for i in range(0,len(price)):
        ws.cell(row = i+1+1,column = 1).font = openpyxl.styles.Font(color=openpyxl.styles.colors.BLUE)
        ws.cell(row = i+1+1, column = 2).number_format = '#,##0.00_);[Green]-#,##0.00'
        ws.cell(row = i+1+1, column = 3).number_format = '#,##0.00_);[Green]-#,##0.00'
        ws.cell(row = i+1+1, column = 4).number_format = '#,##0.00_);[Green]-#,##0.00'
        ws.cell(row = i+1+1,column = 1).value = price[i][0]
        ws.cell(row = i+1+1, column = 2).value = price[i][1][0]
        ws.cell(row = i+1+1, column = 3).value = price[i][1][1]
        ws.cell(row = i+1+1, column = 4).value = price[i][1][2]
    wb.save(today_name_future)
    # if(ALL == False):
    #     tkMessageBox.showinfo(title=u'写入成功', message=u'策略六已经成功，请点击打开进行浏览')

# 7 #波动性:日线 周线 月线10MA都上扬


def future_strategy7():
    synchronization = P.process7()
    wb = open_workbook(today_name_future)
    change_sheet(wb,u'期货代码')
    ws = open_sheet(wb,u'期货代码')
    if len(synchronization):
        last_day = synchronization[0][2][9].isoformat()[0:10]
        this_day = synchronization[0][2][10].isoformat()[0:10]
        last_week = synchronization[0][3][9].isoformat()[0:10]
        this_week = synchronization[0][3][10].isoformat()[0:10]
        last_month = synchronization[0][4][9].isoformat()[0:10]
        this_month = synchronization[0][4][10].isoformat()[0:10]
    else:
        last_day = str(' ')
        this_day = str(' ')
        last_week = str(' ')
        this_week = str(' ')
        last_month = str(' ')
        this_month = str(' ')
    ws.cell(row = 1,column = 3).value = u'日周月10MA'
    ws.cell(row = 1,column = 3).alignment = alignment
    for i in range(len(synchronization)):
        ws.cell(row = i+1+1,column = 3).value = synchronization[i][0]
        ws.cell(row = i+1+1,column = 3).font = openpyxl.styles.Font(color=openpyxl.styles.colors.RED)
    ws = open_sheet(wb,u'日周月10MA')
    ws.cell(row = 1, column = 1).value = u'符合条件代码'
    ws.cell(row = 1, column = 2).value = u'前一天10MA日线\n' + last_day
    ws.cell(row = 1, column = 3).value = u'10MA日线\n' + this_day
    ws.cell(row = 1, column = 4).value = u'前一周10MA周线\n' + last_week
    ws.cell(row = 1, column = 5).value = u'10MA周线\n' + this_week
    ws.cell(row = 1, column = 6).value = u'前一月10MA月线\n' + last_month
    ws.cell(row = 1, column = 7).value = u'10MA月线\n' + this_month
    ws.cell(row=1,column=1).alignment = alignment
    ws.cell(row=1,column=2).alignment = alignment
    ws.cell(row=1,column=3).alignment = alignment
    ws.cell(row=1,column=4).alignment = alignment
    ws.cell(row=1,column=5).alignment = alignment
    ws.cell(row=1,column=6).alignment = alignment
    ws.cell(row=1,column=7).alignment = alignment
    format(ws,7,17.0)
    for i in range(0,len(synchronization)):
        ws.cell(row = i+1+1,column = 1).font = openpyxl.styles.Font(color=openpyxl.styles.colors.BLUE)
        ws.cell(row = i+1+1,column = 1).value = synchronization[i][0]
        ws.cell(row = i+1+1, column = 2).value = synchronization[i][1][0]
        ws.cell(row = i+1+1, column = 3).value = synchronization[i][1][1]
        ws.cell(row = i+1+1, column = 4).value = synchronization[i][1][2]
        ws.cell(row = i+1+1, column = 5).value = synchronization[i][1][3]
        ws.cell(row = i+1+1, column = 6).value = synchronization[i][1][4]
        ws.cell(row = i+1+1, column = 7).value = synchronization[i][1][5]
    wb.save(today_name_future)
    # if(ALL == False):
    #     tkMessageBox.showinfo(title=u'写入成功', message=u'策略七已经成功，请点击打开进行浏览')

# 8 #日线10MA上扬


def future_strategy8():
    DAY10 = P.process8()
    wb = open_workbook(today_name_future)
    change_sheet(wb,u'期货代码')
    ws = open_sheet(wb,u'期货代码')
    ws.cell(row = 1,column = 4).value = u'日10MA'
    ws.cell(row = 1,column = 4).alignment = alignment
    for i in range(len(DAY10)):
        ws.cell(row = i+1+1,column = 4).value = DAY10[i][0]
        ws.cell(row = i+1+1,column = 4).font = openpyxl.styles.Font(color=openpyxl.styles.colors.RED)
    ws = open_sheet(wb,u'日10MA')
    ws.cell(row = 1, column = 1).value = u'符合条件代码'
    ws.cell(row = 1, column = 2).value = u'前一天10MA日线\n' + DAY10[0][2][9].isoformat()[0:10]
    ws.cell(row = 1, column = 3).value = u'10MA日线\n' + DAY10[0][2][10].isoformat()[0:10]
    ws.cell(row=1,column=1).alignment = alignment
    ws.cell(row=1,column=2).alignment = alignment
    ws.cell(row=1,column=3).alignment = alignment
    format(ws,4,17.0)
    for i in range(0,len(DAY10)):
        ws.cell(row = i+1+1,column = 1).font = openpyxl.styles.Font(color=openpyxl.styles.colors.BLUE)
        ws.cell(row = i+1+1,column = 1).value = DAY10[i][0]
        ws.cell(row = i+1+1, column = 2).value = DAY10[i][1][0]
        ws.cell(row = i+1+1, column = 3).value = DAY10[i][1][1]
    wb.save(today_name_future)
    # if(ALL == False):
    #     tkMessageBox.showinfo(title=u'写入成功', message=u'策略八已经成功，请点击打开进行浏览')

# 9 #周线10MA上扬


def future_strategy9():
    WEEK10 = P.process9()
    if len(WEEK10):
        last_day = WEEK10[0][2][9].isoformat()[0:10]
        MA10 = WEEK10[0][2][10].isoformat()[0:10]
    else:
        last_day = str(' ')
        MA10= str(' ')
    wb = open_workbook(today_name_future)
    change_sheet(wb,u'期货代码')
    ws = open_sheet(wb,u'期货代码')
    ws.cell(row = 1,column = 5).value = u'周10MA'
    ws.cell(row = 1,column = 5).alignment = alignment
    for i in range(len(WEEK10)):
        ws.cell(row = i+1+1,column = 5).value = WEEK10[i][0]
        ws.cell(row = i+1+1,column = 5).font = openpyxl.styles.Font(color=openpyxl.styles.colors.RED)
    ws = open_sheet(wb,u'周10MA')
    ws.cell(row = 1, column = 1).value = u'符合条件代码'
    ws.cell(row = 1, column = 2).value = u'前一天10MA周线\n' + last_day
    ws.cell(row = 1, column = 3).value = u'10MA周线\n' + MA10
    ws.cell(row=1,column=1).alignment = alignment
    ws.cell(row=1,column=2).alignment = alignment
    ws.cell(row=1,column=3).alignment = alignment
    format(ws,4,17.0)
    for i in range(0,len(WEEK10)):
        ws.cell(row = i+1+1,column = 1).font = openpyxl.styles.Font(color=openpyxl.styles.colors.BLUE)
        ws.cell(row = i+1+1,column = 1).value = WEEK10[i][0]
        ws.cell(row = i+1+1, column = 2).value = WEEK10[i][1][0]
        ws.cell(row = i+1+1, column = 3).value = WEEK10[i][1][1]
    wb.save(today_name_future)
    # if(ALL == False):
    #     tkMessageBox.showinfo(title=u'写入成功', message=u'策略九已经成功，请点击打开进行浏览')

# 10 #月线10MA上扬


def future_strategy10():
    MON10 = P.process10()

    if len(MON10):
        last_month = MON10[0][2][9].isoformat()[0:10]
        this_month = MON10[0][2][10].isoformat()[0:10]
    else:
        last_month = str(' ')
        this_month = str(' ')
    wb = open_workbook(today_name_future)
    change_sheet(wb,u'期货代码')
    ws = open_sheet(wb,u'期货代码')
    ws.cell(row = 1,column = 6).value = u'月10MA'
    ws.cell(row = 1,column = 6).alignment = alignment
    for i in range(len(MON10)):
        ws.cell(row = i+1+1,column = 6).value = MON10[i][0]
        ws.cell(row = i+1+1,column = 6).font = openpyxl.styles.Font(color=openpyxl.styles.colors.RED)
    ws = open_sheet(wb,u'月10MA')
    ws.cell(row = 1, column = 1).value = u'符合条件代码'
    ws.cell(row = 1, column = 2).value = u'前一天10MA月线\n' + last_month
    ws.cell(row = 1, column = 3).value = u'10MA月线\n' + this_month
    ws.cell(row=1,column=1).alignment = alignment
    ws.cell(row=1,column=2).alignment = alignment
    ws.cell(row=1,column=3).alignment = alignment
    format(ws,4,17.0)
    for i in range(0,len(MON10)):
        ws.cell(row = i+1+1,column = 1).font = openpyxl.styles.Font(color=openpyxl.styles.colors.BLUE)
        ws.cell(row = i+1+1,column = 1).value = MON10[i][0]
        ws.cell(row = i+1+1, column = 2).value = MON10[i][1][0]
        ws.cell(row = i+1+1, column = 3).value = MON10[i][1][1]
    wb.save(today_name_future)
    # if(ALL == False):
    #     tkMessageBox.showinfo(title=u'写入成功', message=u'策略十已经成功，请点击打开进行浏览')

# 11 #60 180 250天波动区间在30% 50% 100%内


def future_strategy11():
    vola_range = P.process11()
    wb = open_workbook(today_name_future)
    change_sheet(wb,u'期货代码')
    ws = open_sheet(wb,u'期货代码')
    ws.cell(row = 1,column = 7).value = u'波动区间内'
    ws.cell(row = 1,column = 7).alignment = alignment
    for i in range(len(vola_range)):
        ws.cell(row = i+1+1,column = 7).value = vola_range[i][0]
        ws.cell(row = i+1+1,column = 7).font = openpyxl.styles.Font(color=openpyxl.styles.colors.RED)
    ws = open_sheet(wb,u'波动区间内')
    ws.cell(row = 1, column = 1).value = u'符合条件代码'
    ws.cell(row = 1, column = 2).value = u'60日内最高价\n('+ vola_range[0][2][0] + u'至今)'
    ws.cell(row = 1, column = 3).value = u'60日内最低价\n('+ vola_range[0][2][0] + u'至今)'
    ws.cell(row = 1, column = 4).value = u'180日内最高价\n('+ vola_range[0][2][1] + u'至今)'
    ws.cell(row = 1, column = 5).value = u'180日内最低价\n('+ vola_range[0][2][1] + u'至今)'
    ws.cell(row = 1, column = 6).value = u'250日内最高价\n('+ vola_range[0][2][2] + u'至今)'
    ws.cell(row = 1, column = 7).value = u'250日内最低价\n('+ vola_range[0][2][2] + u'至今)'
    ws.cell(row=1,column=1).alignment = alignment
    ws.cell(row=1,column=2).alignment = alignment
    ws.cell(row=1,column=3).alignment = alignment
    ws.cell(row=1,column=4).alignment = alignment
    ws.cell(row=1,column=5).alignment = alignment
    ws.cell(row=1,column=6).alignment = alignment
    ws.cell(row=1,column=7).alignment = alignment
    format(ws,7,18.0)
    for i in range(0,len(vola_range)):
        ws.cell(row = i+1+1,column = 1).font = openpyxl.styles.Font(color=openpyxl.styles.colors.BLUE)
        ws.cell(row = i+1+1,column = 1).value = vola_range[i][0]
        ws.cell(row = i+1+1, column = 2).value = vola_range[i][1][0]
        ws.cell(row = i+1+1, column = 3).value = vola_range[i][1][1]
        ws.cell(row = i+1+1, column = 4).value = vola_range[i][1][2]
        ws.cell(row = i+1+1, column = 5).value = vola_range[i][1][3]
        ws.cell(row = i+1+1, column = 6).value = vola_range[i][1][4]
        ws.cell(row = i+1+1, column = 7).value = vola_range[i][1][5]
    wb.save(today_name_future)
    # if(ALL == False):
    #     tkMessageBox.showinfo(title=u'写入成功', message=u'策略十一已经成功，请点击打开进行浏览')

# 12 #60天波动区间在30%内


def future_strategy12():
    vola_range60 = P.process12()
    wb = open_workbook(today_name_future)
    change_sheet(wb,u'期货代码')
    ws = open_sheet(wb,u'期货代码')
    ws.cell(row = 1,column = 8).value = u'60日波动小于30%'
    ws.cell(row = 1,column = 8).alignment = alignment
    for i in range(len(vola_range60)):
        ws.cell(row = i+1+1,column = 8).value = vola_range60[i][0]
        ws.cell(row = i+1+1,column = 8).font = openpyxl.styles.Font(color=openpyxl.styles.colors.RED)
    ws = open_sheet(wb,u'60日波动小于30%')
    ws.cell(row = 1, column = 1).value = u'符合条件代码'
    ws.cell(row = 1, column = 2).value = u'60日内最高价\n（'+ vola_range60[0][2] + u'至今)'
    ws.cell(row = 1, column = 3).value = u'60日内最低价\n（'+ vola_range60[0][2] + u'至今)'
    ws.cell(row=1,column=1).alignment = alignment
    ws.cell(row=1,column=2).alignment = alignment
    ws.cell(row=1,column=3).alignment = alignment
    format(ws,4,18.0)
    for i in range(0,len(vola_range60)):
        ws.cell(row = i+1+1,column = 1).font = openpyxl.styles.Font(color=openpyxl.styles.colors.BLUE)
        ws.cell(row = i+1+1,column = 1).value = vola_range60[i][0]
        ws.cell(row = i+1+1, column = 2).value = vola_range60[i][1][0]
        ws.cell(row = i+1+1, column = 3).value = vola_range60[i][1][1]
    wb.save(today_name_future)
    # if(ALL == False):
    #     tkMessageBox.showinfo(title=u'写入成功', message=u'策略十二已经成功，请点击打开进行浏览')

# 13 #180天波动区间在50%内


def future_strategy13():
    vola_range180 = P.process13()
    wb = open_workbook(today_name_future)
    change_sheet(wb,u'期货代码')
    ws = open_sheet(wb,u'期货代码')
    ws.cell(row = 1,column = 9).value = u'180日波动小于50%'
    ws.cell(row = 1,column = 9).alignment = alignment
    for i in range(len(vola_range180)):
        ws.cell(row = i+1+1,column = 9).value = vola_range180[i][0]
        ws.cell(row = i+1+1,column = 9).font = openpyxl.styles.Font(color=openpyxl.styles.colors.RED)
    ws = open_sheet(wb,u'180日波动小于50%')
    ws.cell(row = 1, column = 1).value = u'符合条件代码'
    ws.cell(row = 1, column = 2).value = u'180日内最高价\n（'+ vola_range180[0][2] + u'至今）'
    ws.cell(row = 1, column = 3).value = u'180日内最低价\n（'+ vola_range180[0][2] + u'至今）'
    ws.cell(row=1,column=1).alignment = alignment
    ws.cell(row=1,column=2).alignment = alignment
    ws.cell(row=1,column=3).alignment = alignment
    format(ws,4,18.0)
    for i in range(0,len(vola_range180)):
        ws.cell(row = i+1+1,column = 1).font = openpyxl.styles.Font(color=openpyxl.styles.colors.BLUE)
        ws.cell(row = i+1+1,column = 1).value = vola_range180[i][0]
        ws.cell(row = i+1+1, column = 2).value = vola_range180[i][1][0]
        ws.cell(row = i+1+1, column = 3).value = vola_range180[i][1][1]
    wb.save(today_name_future)
    # if(ALL == False):
    #     tkMessageBox.showinfo(title=u'写入成功', message=u'策略十三已经成功，请点击打开进行浏览')

# 14 #250天波动区间在100%内


def future_strategy14():
    vola_range250 = P.process14()
    wb = open_workbook(today_name_future)
    change_sheet(wb,u'期货代码')
    ws = open_sheet(wb,u'期货代码')
    ws.cell(row = 1,column = 10).value = u'250日波动小于100%'
    ws.cell(row = 1,column = 10).alignment = alignment
    for i in range(len(vola_range250)):
        ws.cell(row = i+1+1,column = 10).value = vola_range250[i][0]
        ws.cell(row = i+1+1,column = 10).font = openpyxl.styles.Font(color=openpyxl.styles.colors.RED)
    ws = open_sheet(wb,u'250日波动小于100%')
    ws.cell(row = 1, column = 1).value = u'符合条件代码'
    ws.cell(row = 1, column = 2).value = u'250日内最高价\n（'+ vola_range250[0][2] + u'至今）'
    ws.cell(row = 1, column = 3).value = u'250日内最低价\n（'+ vola_range250[0][2] + u'至今）'
    ws.cell(row=1,column=1).alignment = alignment
    ws.cell(row=1,column=2).alignment = alignment
    ws.cell(row=1,column=3).alignment = alignment
    format(ws,4,18.0)
    for i in range(0,len(vola_range250)):
        ws.cell(row = i+1+1,column = 1).font = openpyxl.styles.Font(color=openpyxl.styles.colors.BLUE)
        ws.cell(row = i+1+1,column = 1).value = vola_range250[i][0]
        ws.cell(row = i+1+1, column = 2).value = vola_range250[i][1][0]
        ws.cell(row = i+1+1, column = 3).value = vola_range250[i][1][1]
    wb.save(today_name_future)
    # if(ALL == False):
    #     tkMessageBox.showinfo(title=u'写入成功', message=u'策略十四已经成功，请点击打开进行浏览')
# 24 #涨幅日、周、月都超过大盘


def future_strategy24():
    increase = P.process24()
    if len(increase):
        last_one = increase[0][2][19].isoformat()[0:10]
        last_five = increase[0][2][15].isoformat()[0:10]
        last_twenty = increase[0][2][0].isoformat()[0:10]
    else:
        last_one = str(' ')
        last_five = str(' ')
        last_twenty = str(' ')
    wb = open_workbook(today_name_future)
    change_sheet(wb,u'期货代码')
    ws = open_sheet(wb,u'期货代码')
    ws.cell(row = 1,column = 11).value = u'涨幅'
    ws.cell(row = 1,column = 11).alignment = alignment
    for i in range(len(increase)):
        ws.cell(row=i + 1 + 1, column=11).value = increase[i][0]
        ws.cell(row=i + 1 + 1, column=11).font = openpyxl.styles.Font(color=openpyxl.styles.colors.RED)
    ws = open_sheet(wb, u'涨幅')
    ws.cell(row = 1, column = 1).value = u'符合条件代码'
    ws.cell(row = 1, column = 2).value = u'日涨幅\n（前一天）\n' + last_one
    ws.cell(row = 1, column = 3).value = u'周涨幅\n（前五天）\n' + last_five
    ws.cell(row = 1, column = 4).value = u'月涨幅\n（二十天）\n' + last_twenty
    ws.cell(row = 1, column = 5).value = u'大盘日涨幅\n（前一天）\n' + last_one
    ws.cell(row = 1, column = 6).value = u'大盘周涨幅\n（前五天）\n' + last_five
    ws.cell(row = 1, column = 7).value = u'大盘月涨幅\n（二十天）\n' + last_twenty
    ws.cell(row=1,column=1).alignment = alignment
    ws.cell(row=1,column=2).alignment = alignment
    ws.cell(row=1,column=3).alignment = alignment
    ws.cell(row=1,column=4).alignment = alignment
    ws.cell(row=1,column=5).alignment = alignment
    ws.cell(row=1,column=6).alignment = alignment
    ws.cell(row=1,column=7).alignment = alignment
    format(ws, 7, 15.0)
    for i in range(0,len(increase)):
        ws.cell(row = i+1+1,column = 1).font = openpyxl.styles.Font(color=openpyxl.styles.colors.BLUE)
        ws.cell(row = i+1+1,column = 2).font = openpyxl.styles.Font(color=openpyxl.styles.colors.RED)
        ws.cell(row = i+1+1,column = 3).font = openpyxl.styles.Font(color=openpyxl.styles.colors.RED)
        ws.cell(row = i+1+1,column = 4).font = openpyxl.styles.Font(color=openpyxl.styles.colors.RED)
        ws.cell(row = i+1+1, column = 2).number_format = '#,##0.00%_);[Green]-#,##0.00%'
        ws.cell(row = i+1+1, column = 3).number_format = '#,##0.00%_);[Green]-#,##0.00%'
        ws.cell(row = i+1+1, column = 4).number_format = '#,##0.00%_);[Green]-#,##0.00%'
        ws.cell(row = i+1+1, column = 5).number_format = '#,##0.00%_);[Green]-#,##0.00%'
        ws.cell(row = i+1+1, column = 6).number_format = '#,##0.00%_);[Green]-#,##0.00%'
        ws.cell(row = i+1+1, column = 7).number_format = '#,##0.00%_);[Green]-#,##0.00%'
        ws.cell(row = i+1+1,column = 1).value = increase[i][0]
        ws.cell(row = i+1+1, column = 2).value = increase[i][1][0]
        ws.cell(row = i+1+1, column = 3).value = increase[i][1][1]
        ws.cell(row = i+1+1, column = 4).value = increase[i][1][2]
        ws.cell(row = i+1+1, column = 5).value = increase[i][1][3]
        ws.cell(row = i+1+1, column = 6).value = increase[i][1][4]
        ws.cell(row = i+1+1, column = 7).value = increase[i][1][4]
    wb.save(today_name_future)
# 25 #袁氏选股


def future_strategy25():
    yuan1,yuan2,all = P.process25()
    if len(yuan1):
        day_today = yuan1[0][2][252].isoformat()[0:10]
        day_pre_5 = yuan1[0][2][247].isoformat()[0:10]
        day_pre_60 = yuan1[0][2][192].isoformat()[0:10]
    else:
        day_today = str(' ')
        day_pre_5 = str(' ')
        day_pre_60 = str(' ')
    if len(yuan2):
        day_today = yuan2[0][2][252].isoformat()[0:10]
        day_pre_5 = yuan2[0][2][247].isoformat()[0:10]
        dau_pre_252 = yuan2[0][2][0].isoformat()[0:10]
    else:
        day_today = str(' ')
        day_pre_5 = str(' ')
        dau_pre_252 = str(' ')
    if len(all):
        day_today = yuan1[0][2][252].isoformat()[0:10]
        day_pre_5 = yuan1[0][2][247].isoformat()[0:10]
        day_pre_60 = yuan1[0][2][192].isoformat()[0:10]
        dau_pre_252 = yuan1[0][2][0].isoformat()[0:10]
    else:
        day_today = str(' ')
        day_pre_5 = str(' ')
        day_pre_60 = str(' ')
        dau_pre_252 = str(' ')
    wb = open_workbook(today_name_future)
    change_sheet(wb, u'期货代码')
    ws = open_sheet(wb, u'期货代码')
    ws.cell(row = 1,column = 12).value = u'袁氏选股\n(积极版)'
    ws.cell(row = 1,column = 12).alignment = alignment
    ws.cell(row = 1,column = 13).value = u'袁氏选股\n(年版)'
    ws.cell(row = 1,column = 13).alignment = alignment
    for i in range(len(yuan1)):
        ws.cell(row=i + 1 + 1, column=12).value = yuan1[i][0]
        ws.cell(row=i + 1 + 1, column=12).font = openpyxl.styles.Font(color=openpyxl.styles.colors.RED)
    for i in range(len(yuan2)):
        ws.cell(row=i + 1 + 1, column=13).value = yuan2[i][0]
        ws.cell(row=i + 1 + 1, column=13).font = openpyxl.styles.Font(color=openpyxl.styles.colors.RED)
    ws = open_sheet(wb, u'袁氏选股')
    ws.cell(row=1, column=1).value = u'符合条件代码（积极版）'
    ws.cell(row = 1, column = 2).value = u'收盘价\n（当日）\n' + day_today
    ws.cell(row = 1, column = 3).value = u'收盘价\n（前五天）\n' + day_pre_5
    ws.cell(row = 1, column = 4).value = u'收盘价\n（前六十天）\n' + day_pre_60
    ws.cell(row=1, column=5).value = u'符合条件代码（年版）'
    ws.cell(row = 1, column = 6).value = u'收盘价\n（当日）\n' + day_today
    ws.cell(row = 1, column = 7).value = u'收盘价\n（前五天）\n' + day_pre_5
    ws.cell(row = 1, column = 8).value = u'收盘价\n（前252天）\n' + dau_pre_252
    ws.cell(row = 1, column = 9).value = u'全部代码'
    ws.cell(row = 1, column = 10).value = u'收盘价\n（当日）\n' + day_today
    ws.cell(row = 1, column = 11).value = u'收盘价\n（前五天）\n' + day_pre_5
    ws.cell(row = 1, column = 12).value = u'收盘价\n（前六十天）\n' + day_pre_60
    ws.cell(row = 1, column = 13).value = u'收盘价\n（前252天）\n' + dau_pre_252
    ws.cell(row=1,column=1).alignment = alignment
    ws.cell(row=1,column=2).alignment = alignment
    ws.cell(row=1,column=3).alignment = alignment
    ws.cell(row=1,column=4).alignment = alignment
    ws.cell(row=1,column=5).alignment = alignment
    ws.cell(row=1,column=6).alignment = alignment
    ws.cell(row=1,column=7).alignment = alignment
    ws.cell(row=1,column=8).alignment = alignment
    ws.cell(row=1,column=9).alignment = alignment
    ws.cell(row=1,column=10).alignment = alignment
    ws.cell(row=1,column=11).alignment = alignment
    ws.cell(row=1,column=12).alignment = alignment
    ws.cell(row=1,column=13).alignment = alignment
    ws.column_dimensions['A'].width = 15.0
    format(ws, 13, 15.0)
    for i in range(0, len(yuan1)):
        ws.cell(row=1, column=1).alignment = alignment
        ws.cell(row = i+1+1,column = 1).font = openpyxl.styles.Font(color=openpyxl.styles.colors.BLUE)
        ws.cell(row = i+1+1,column = 2).font = openpyxl.styles.Font(color=openpyxl.styles.colors.RED)
        ws.cell(row = i+1+1,column = 3).font = openpyxl.styles.Font(color=openpyxl.styles.colors.RED)
        ws.cell(row = i+1+1,column = 4).font = openpyxl.styles.Font(color=openpyxl.styles.colors.RED)
        ws.cell(row = i+1+1,column = 1).value = yuan1[i][0]
        ws.cell(row = i+1+1, column = 2).value = yuan1[i][1][0]
        ws.cell(row = i+1+1, column = 3).value = yuan1[i][1][1]
        ws.cell(row = i+1+1, column = 4).value = yuan1[i][1][2]
    for i in range(0, len(yuan2)):
        ws.cell(row=1, column=5).alignment = alignment
        ws.cell(row = i+1+1,column = 5).font = openpyxl.styles.Font(color=openpyxl.styles.colors.BLUE)
        ws.cell(row = i+1+1,column = 6).font = openpyxl.styles.Font(color=openpyxl.styles.colors.RED)
        ws.cell(row = i+1+1,column = 7).font = openpyxl.styles.Font(color=openpyxl.styles.colors.RED)
        ws.cell(row = i+1+1,column = 8).font = openpyxl.styles.Font(color=openpyxl.styles.colors.RED)
        ws.cell(row = i+1+1, column = 5).value = yuan2[i][0]
        ws.cell(row = i+1+1, column = 6).value = yuan2[i][1][0]
        ws.cell(row = i+1+1, column = 7).value = yuan2[i][1][1]
        ws.cell(row = i+1+1, column = 8).value = yuan2[i][1][2]
    for i in range(0,len(all)):
        ws.cell(row=1, column=9).alignment = alignment
        ws.cell(row=i + 1 + 1, column=9).font = openpyxl.styles.Font(color=openpyxl.styles.colors.BLUE)
        ws.cell(row=i + 1 + 1, column=10).font = openpyxl.styles.Font(color=openpyxl.styles.colors.RED)
        ws.cell(row=i + 1 + 1, column=11).font = openpyxl.styles.Font(color=openpyxl.styles.colors.RED)
        ws.cell(row=i + 1 + 1, column=12).font = openpyxl.styles.Font(color=openpyxl.styles.colors.RED)
        ws.cell(row=i + 1 + 1, column=13).font = openpyxl.styles.Font(color=openpyxl.styles.colors.RED)
        ws.cell(row = i+1+1, column = 9).value = all[i][0]
        ws.cell(row = i+1+1, column = 10).value = all[i][1]
        ws.cell(row = i+1+1, column = 11).value = all[i][2]
        ws.cell(row = i+1+1, column = 12).value = all[i][3]
        ws.cell(row = i+1+1, column = 13).value = all[i][4]

    wb.save(today_name_future)
#执行所有策略
def all_strategy():
    # tkMessageBox.showinfo(title=u'开始执行', message=u'执行所有策略耗时较长，点击确定开始，请耐心等待')
    write_all()
    strategy1()
    strategy2()
    strategy3()
    strategy4()
    strategy5()
    strategy6()
    strategy7()
    strategy8()
    strategy9()
    strategy10()
    strategy11()
    strategy12()
    strategy13()
    strategy14()
    strategy15()
    strategy16()
    strategy17()
    strategy18()
    strategy19()
    strategy202122()
    strategy23()
    strategy24()
    strategy25()
    strategy27()
    strategy29()
    strategy30()
    strategy31()
    star()
    # tkMessageBox.showinfo(title=u'写入成功', message=u'所有策略已经成功，请点击打开进行浏览')

#为期货执行策略
def all_strategy_future():
    D.Codes = D.future
    future_strategy5()
    future_strategy6()
    future_strategy7()
    future_strategy8()
    future_strategy9()
    future_strategy10()
    future_strategy11()
    future_strategy12()
    future_strategy13()
    future_strategy14()
    future_strategy24()
    future_strategy25()